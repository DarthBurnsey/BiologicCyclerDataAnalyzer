"""Tests for the NMC cathode workbook ontology import service."""

from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.database import Base
from app.models.ontology import CellBuild, ElectrodeBatch, LineageEdge, Material
from app.services.nmc_cathode_workbook import _normalize_legacy_formulation, import_nmc_cathode_workbook


pytestmark = pytest.mark.asyncio


SUMMARY_HEADERS = [
    "Date",
    "Test ",
    "NMC811",
    "PVDF",
    "Cx",
    "Super P/Carbon Black",
    "CNTs",
    "Solids Content",
    "Mixing Strategy",
    "CC",
    "Electrolyte",
    "1st Dis. Cap. (mAh/g)",
    "Reversible Capacity (mAh/g)",
    "FCE (%)",
    "Cycle Life ",
    "Cycler ",
    "Cycling Strategy ",
    "Results",
    "Notes",
]


def create_session_factory(tmp_path: Path):
    db_path = tmp_path / "ontology-import-test.db"
    engine = create_async_engine(f"sqlite+aiosqlite:///{db_path}")
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    return engine, session_factory


async def initialize_backend_db(engine) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Experimental Data"
    for index, header in enumerate(SUMMARY_HEADERS, start=1):
        summary.cell(1, index).value = header

    summary.append(
        [
            datetime(2025, 9, 11),
            "N1",
            0.94,
            0.03,
            0.03,
            0.0,
            0.0,
            0.4548,
            "Arkema Protocol w/ media",
            "Al",
            "1M LiTFSI 3:7 +10% FEC",
            None,
            None,
            None,
            None,
            "B1",
            "3x Form., Staircase",
            None,
            "Initial testing of NMC811 (MSE Supplies).",
        ]
    )
    summary.append(
        [
            datetime(2025, 9, 16),
            "N1B",
            0.94,
            0.03,
            0.03,
            0.0,
            0.0,
            0.4548,
            "Arkema Protocol w/ media",
            "Al",
            "LiPF6 (1:1:1)",
            None,
            None,
            None,
            None,
            "B2",
            "3x Form., Staircase",
            None,
            "Variant test with different electrolyte. | Tracking: cal",
        ]
    )
    summary.append(
        [
            datetime(2025, 9, 20),
            "N3",
            0.95,
            0.025,
            0.025,
            0.0,
            0.0,
            0.45,
            "Arkema Protocol w/ media",
            "Al",
            "1M LiPF6 (1:1:1)",
            None,
            None,
            None,
            None,
            "B4",
            "3x Form., Staircase",
            None,
            "Parent-only batch.",
        ]
    )

    workbook.create_sheet("Density")

    yellow = PatternFill(fill_type="solid", fgColor="FFFFFF00")

    n1 = workbook.create_sheet("N1")
    n1["A1"] = "Experiment"
    n1["B1"] = "N1"
    n1["C1"] = "Notes"
    n1["D1"] = "Begin testing NMC formulations."
    n1["E1"] = "Result Notes"
    n1["F1"] = "Electrode quality looks good."
    n1["A2"] = "Date"
    n1["B2"] = datetime(2025, 9, 11)
    n1["A4"] = "NMC811"
    n1["B4"] = 0.94
    n1["A5"] = "Cx (echo)"
    n1["B5"] = 0.03
    n1["A8"] = "PVDF HSV1810"
    n1["B8"] = 0.03
    n1["D23"] = "Electrode Thickness (um) "
    n1["E23"] = "Pressed Thickness (um) "
    n1["D24"] = 84
    n1["E24"] = 70
    n1["E17"] = "Disc Mass (mg)"
    n1["E18"] = 25.43

    n16 = workbook.create_sheet("N16")
    n16["A1"] = "Experiment"
    n16["B1"] = "N16"
    n16["C1"] = "Notes"
    n16["D1"] = "Testing of Hx-e (N16a:HCl conc. acid washed and N16b: HCl+Base washed- Eli)"
    n16["A2"] = "Date"
    n16["B2"] = datetime(2026, 3, 6)
    n16["A4"] = "NMC811"
    n16["B4"] = 0.95
    n16["A5"] = "Hx-e (HCl Acid Wash)"
    n16["B5"] = 0.025
    n16["A8"] = "PVDF HSV1810"
    n16["B8"] = 0.025
    n16["B29"] = "Disc Mass (mg)"
    n16["C29"] = 35.34
    n16["C29"].fill = yellow
    n16["B30"] = "Electrode Thickness (um) "
    n16["C30"] = 95
    n16["C30"].fill = yellow
    n16["B31"] = "Pressed Thickness (um) "
    n16["C31"] = 64
    n16["C31"].fill = yellow

    n3 = workbook.create_sheet("N3")
    n3["A1"] = "Experiment"
    n3["B1"] = "N3"
    n3["C1"] = "Notes"
    n3["D1"] = "Parent-only batch for ontology lineage."
    n3["A2"] = "Date"
    n3["B2"] = datetime(2025, 9, 20)
    n3["A4"] = "NMC811"
    n3["B4"] = 0.95
    n3["A5"] = "Cx (echo)"
    n3["B5"] = 0.025
    n3["A8"] = "PVDF HSV1810"
    n3["B8"] = 0.025

    workbook.save(path)


def build_legacy_db(path: Path) -> None:
    conn = sqlite3.connect(path)
    try:
        conn.executescript(
            """
            CREATE TABLE projects (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                project_type TEXT
            );

            CREATE TABLE cell_experiments (
                id INTEGER PRIMARY KEY,
                project_id INTEGER NOT NULL,
                cell_name TEXT NOT NULL,
                loading REAL,
                active_material REAL,
                electrolyte TEXT,
                formulation_json TEXT,
                solids_content REAL,
                pressed_thickness REAL,
                experiment_notes TEXT,
                substrate TEXT,
                separator TEXT,
                data_json TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO projects (id, name, project_type) VALUES (1, 'NMC Half Cells', 'Cathode')"
        )
        conn.execute(
            "INSERT INTO projects (id, name, project_type) VALUES (2, 'NMC- Si Full Cell', 'Full Cell')"
        )

        n1_data = {
            "tracking": {"tracking_date": "2025-09-12"},
            "cells": [
                {
                    "cell_name": "N1 i",
                    "test_number": "N1 i",
                    "electrolyte": "1M LiPF6 (1:1:1)",
                    "separator": "2+12+2 CCS",
                    "substrate": "Aluminum",
                    "loading": 25.43,
                    "active_material": 94.0,
                }
            ]
        }
        conn.execute(
            """
            INSERT INTO cell_experiments (
                id, project_id, cell_name, loading, active_material, electrolyte,
                formulation_json, solids_content, pressed_thickness, experiment_notes,
                substrate, separator, data_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                10,
                1,
                "N1",
                25.43,
                94.0,
                "1M LiPF6 (1:1:1)",
                json.dumps(
                    [
                        {"Component": "NMC811", "Dry Mass Fraction (%)": 94.0},
                        {"Component": "PVDF HSV1810", "Dry Mass Fraction (%)": 3.0},
                        {"Component": "Hx-T", "Dry Mass Fraction (%)": 3.0},
                    ]
                ),
                45.48,
                70.0,
                "Legacy exact parent row.",
                "Aluminum",
                "2+12+2 CCS",
                json.dumps(n1_data),
            ),
        )

        n1b_data = {
            "cells": [
                {
                    "cell_name": "N1b i",
                    "test_number": "N1b i",
                    "electrolyte": "LiPF6 (1:1:1)",
                    "separator": "2+12+2 CCS",
                    "substrate": "Aluminum",
                    "loading": 25.43,
                    "active_material": 94.0,
                }
            ]
        }
        conn.execute(
            """
            INSERT INTO cell_experiments (
                id, project_id, cell_name, loading, active_material, electrolyte,
                formulation_json, solids_content, pressed_thickness, experiment_notes,
                substrate, separator, data_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                11,
                1,
                "N1b",
                25.43,
                94.0,
                "LiPF6 (1:1:1)",
                json.dumps(
                    [
                        {"Component": "NMC811", "Dry Mass Fraction (%)": 94.0},
                        {"Component": "PVDF HSV1810", "Dry Mass Fraction (%)": 3.0},
                        {"Component": "Hx-T", "Dry Mass Fraction (%)": 3.0},
                    ]
                ),
                45.48,
                70.0,
                "Variant row.",
                "Aluminum",
                "2+12+2 CCS",
                json.dumps(n1b_data),
            ),
        )

        n16a_data = {
            "cells": [
                {
                    "cell_name": "N16a i",
                    "test_number": "N16a i",
                    "electrolyte": "1M LiPF6 EC:DMC:EMC (1:1:1) + 2% VC + 5% FEC",
                    "separator": "2+12+2 CCS",
                    "substrate": "Aluminum",
                    "loading": 12.1,
                    "active_material": 95.0,
                }
            ]
        }
        conn.execute(
            """
            INSERT INTO cell_experiments (
                id, project_id, cell_name, loading, active_material, electrolyte,
                formulation_json, solids_content, pressed_thickness, experiment_notes,
                substrate, separator, data_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                20,
                1,
                "N16a",
                12.1,
                95.0,
                "1M LiPF6 EC:DMC:EMC (1:1:1) + 2% VC + 5% FEC",
                json.dumps(
                    [
                        {"Component": "NMC811 (BASF)", "Dry Mass Fraction (%)": 95.0},
                        {"Component": "PVDF HSV1810", "Dry Mass Fraction (%)": 2.5},
                        {"Component": "Hx-e (HCl Acid)", "Dry Mass Fraction (%)": 2.5},
                    ]
                ),
                57.7,
                64.0,
                "Legacy-only branch.",
                "Aluminum",
                "2+12+2 CCS",
                json.dumps(n16a_data),
            ),
        )

        n1a_data = {
            "cells": [
                {
                    "cell_name": "N1a i",
                    "test_number": "N1a i",
                    "electrolyte": "1M LiTFSI 3:7 +10% FEC",
                    "separator": "2+12+2 CCS",
                    "substrate": "Aluminum",
                    "loading": 25.43,
                    "active_material": 94.0,
                }
            ]
        }
        conn.execute(
            """
            INSERT INTO cell_experiments (
                id, project_id, cell_name, loading, active_material, electrolyte,
                formulation_json, solids_content, pressed_thickness, experiment_notes,
                substrate, separator, data_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                21,
                1,
                "N1a",
                25.43,
                94.0,
                "1M LiTFSI 3:7 +10% FEC",
                json.dumps(
                    [
                        {"Component": "NMC811", "Dry Mass Fraction (%)": 94.0},
                        {"Component": "PVDF HSV1810", "Dry Mass Fraction (%)": 3.0},
                        {"Component": "Hx-T", "Dry Mass Fraction (%)": 3.0},
                    ]
                ),
                45.48,
                70.0,
                "Legacy-only branch for electrolyte evaluation.",
                "Aluminum",
                "2+12+2 CCS",
                json.dumps(n1a_data),
            ),
        )

        conn.commit()
    finally:
        conn.close()


def build_comparison_workbook(path: Path) -> None:
    workbook = Workbook()
    compiled = workbook.active
    compiled.title = "Compiled"
    compiled.append(
        [
            "Experiment",
            "ldg",
            "AM",
            "1st Chg",
            "1st Dis",
            "Sp.Dis (mAh/g)",
            "FCE",
            "Electrolytes",
            "Purpose",
            "Notes",
        ]
    )
    compiled.append(
        [
            "N1",
            25.43,
            23.91,
            None,
            None,
            180.0,
            0.91,
            "1M LiPF6 (1:1:1)",
            "Override primary workbook electrolyte",
            "Comparison workbook should win for electrolyte",
        ]
    )
    summary = workbook.create_sheet("Summary")
    summary.append(
        ["Experiment", "Spec. Dis (mAh/g)", "FCE %", "Carbon Material", "Electrolytes", "Notes", "Purpose"]
    )
    summary.append(
        ["N1", 180.0, 0.91, "Hx-T", "1M LiPF6 (1:1:1)", "Uncal", "Override primary workbook electrolyte"]
    )
    workbook.save(path)


def append_n12_loading_backfill_case(workbook_path: Path, legacy_db_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    summary = workbook["Experimental Data"]
    summary.append(
        [
            datetime(2026, 1, 7),
            "N12",
            0.94,
            0.03,
            0.03,
            0.0,
            0.0,
            0.5,
            "Arkema Protocol w/ media",
            "Al",
            "1M LiPF6 (1:1:1)",
            None,
            None,
            None,
            None,
            "B6",
            "Staircase",
            None,
            "Parent with highlighted loading that should backfill from legacy child cells.",
        ]
    )
    yellow = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    n12 = workbook.create_sheet("N12")
    n12["A1"] = "Experiment"
    n12["B1"] = "N12"
    n12["D1"] = "Parent N12 sheet"
    n12["A2"] = "Date"
    n12["B2"] = datetime(2026, 1, 7)
    n12["A4"] = "NMC811"
    n12["B4"] = 0.94
    n12["A5"] = "Hx-e"
    n12["B5"] = 0.03
    n12["A8"] = "PVDF HSV1810"
    n12["B8"] = 0.03
    n12["B29"] = "Disc Mass (mg)"
    n12["C29"] = 13.6
    n12["C29"].fill = yellow
    workbook.save(workbook_path)

    conn = sqlite3.connect(legacy_db_path)
    try:
        n12b_data = {
            "cells": [
                {
                    "cell_name": "N12b i",
                    "test_number": "N12b i",
                    "electrolyte": "1M LiPF6 (1:1:1)",
                    "separator": "2+12+2 CCS",
                    "substrate": "Aluminum",
                    "loading": 13.6,
                    "active_material": 94.0,
                },
                {
                    "cell_name": "N12b ii",
                    "test_number": "N12b ii",
                    "electrolyte": "1M LiPF6 (1:1:1)",
                    "separator": "2+12+2 CCS",
                    "substrate": "Aluminum",
                    "loading": 13.6,
                    "active_material": 94.0,
                },
            ]
        }
        conn.execute(
            """
            INSERT INTO cell_experiments (
                id, project_id, cell_name, loading, active_material, electrolyte,
                formulation_json, solids_content, pressed_thickness, experiment_notes,
                substrate, separator, data_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                30,
                1,
                "N12b",
                13.6,
                94.0,
                "1M LiPF6 (1:1:1)",
                json.dumps(
                    [
                        {"Component": "NMC811 (BASF)", "Dry Mass Fraction (%)": 94.0},
                        {"Component": "PVDF HSV1810", "Dry Mass Fraction (%)": 3.0},
                        {"Component": "Hx-e", "Dry Mass Fraction (%)": 3.0},
                    ]
                ),
                50.0,
                49.0,
                "Child branch for higher loading study.",
                "Aluminum",
                "2+12+2 CCS",
                json.dumps(n12b_data),
            ),
        )
        conn.commit()
    finally:
        conn.close()


async def test_preview_import_reports_review_items_and_mismatches(tmp_path):
    workbook_path = tmp_path / "NMC Cathode Formulations 2025.xlsx"
    legacy_db_path = tmp_path / "cellscope.db"
    build_workbook(workbook_path)
    build_legacy_db(legacy_db_path)

    engine, session_factory = create_session_factory(tmp_path)
    await initialize_backend_db(engine)

    async with session_factory() as session:
        preview = await import_nmc_cathode_workbook(
            session,
            workbook_path=str(workbook_path),
            legacy_db_path=str(legacy_db_path),
            persist=False,
        )

    assert preview["persisted"] is False
    assert preview["counts"]["parent_batches"] == 2
    assert preview["counts"]["variant_batches"] >= 3
    assert any(item["name"] == "Hx-T" for item in preview["materials"])
    assert any(item["record_name"] == "N16" and item["field"] == "pressed_thickness_um" for item in preview["review_items"])
    assert any(item["record_name"] == "N1" and item["field"] == "electrolyte" for item in preview["mismatch_items"])
    assert not any(item["record_name"] == "N3" and item["status"] == "missing_in_legacy" for item in preview["mismatch_items"])
    assert not any(item["record_name"] == "N1a" and item["status"] == "legacy_only" for item in preview["mismatch_items"])
    n16a_variant = next(item for item in preview["variant_batches"] if item["batch_name"] == "N16a")
    assert not any(item["batch_name"] == "N16" for item in preview["parent_batches"])
    assert n16a_variant["parent_batch_name"] is None
    assert n16a_variant["source_kind"] == "independent_root_variant"
    hx_e_component = next(component for component in n16a_variant["formulation_json"] if component["Component"] == "Hx-e")
    assert hx_e_component["metadata"]["post_processing"] == "concentrated_hcl_acid_wash"
    n1a_variant = next(item for item in preview["variant_batches"] if item["batch_name"] == "N1a")
    assert n1a_variant["metadata_json"]["override_metadata"]["study_focus"] == "electrolyte_evaluation"

    await engine.dispose()


async def test_comparison_workbook_overrides_primary_workbook_rows(tmp_path):
    workbook_path = tmp_path / "NMC Cathode Formulations 2025.xlsx"
    comparison_workbook_path = tmp_path / "NMC Cathode Data Comparison N4 onwards.xlsx"
    legacy_db_path = tmp_path / "cellscope.db"
    build_workbook(workbook_path)
    build_comparison_workbook(comparison_workbook_path)
    build_legacy_db(legacy_db_path)

    engine, session_factory = create_session_factory(tmp_path)
    await initialize_backend_db(engine)

    async with session_factory() as session:
        preview = await import_nmc_cathode_workbook(
            session,
            workbook_path=str(workbook_path),
            legacy_db_path=str(legacy_db_path),
            comparison_workbook_path=str(comparison_workbook_path),
            persist=False,
        )

    assert not any(item["record_name"] == "N1" and item["field"] == "electrolyte" for item in preview["mismatch_items"])
    n1_parent = next(item for item in preview["parent_batches"] if item["batch_name"] == "N1")
    assert n1_parent["metadata_json"]["default_electrolyte"] == "1M LiPF6 (1:1:1)"
    assert n1_parent["metadata_json"]["comparison_ref"] == "Summary!2"

    await engine.dispose()


async def test_export_dir_writes_reconciliation_artifacts(tmp_path):
    workbook_path = tmp_path / "NMC Cathode Formulations 2025.xlsx"
    legacy_db_path = tmp_path / "cellscope.db"
    export_dir = tmp_path / "reconciliation"
    build_workbook(workbook_path)
    build_legacy_db(legacy_db_path)

    engine, session_factory = create_session_factory(tmp_path)
    await initialize_backend_db(engine)

    async with session_factory() as session:
        preview = await import_nmc_cathode_workbook(
            session,
            workbook_path=str(workbook_path),
            legacy_db_path=str(legacy_db_path),
            export_dir=str(export_dir),
            persist=False,
        )

    review_export = preview["review_export"]
    assert Path(review_export["export_dir"]).exists()
    assert Path(review_export["preview_json"]).exists()
    assert Path(review_export["issues_csv"]).exists()
    assert Path(review_export["readme_md"]).exists()
    assert "NMC Cathode Ontology Reconciliation" in Path(review_export["readme_md"]).read_text(encoding="utf-8")
    csv_text = Path(review_export["issues_csv"]).read_text(encoding="utf-8")
    assert "record_name" in csv_text
    assert "N16" in csv_text

    await engine.dispose()


async def test_persist_import_creates_batches_lineage_and_cell_builds(tmp_path):
    workbook_path = tmp_path / "NMC Cathode Formulations 2025.xlsx"
    legacy_db_path = tmp_path / "cellscope.db"
    build_workbook(workbook_path)
    build_legacy_db(legacy_db_path)

    engine, session_factory = create_session_factory(tmp_path)
    await initialize_backend_db(engine)

    async with session_factory() as session:
        result = await import_nmc_cathode_workbook(
            session,
            workbook_path=str(workbook_path),
            legacy_db_path=str(legacy_db_path),
            persist=True,
        )
        await session.commit()

    assert result["persisted"] is True

    async with session_factory() as session:
        materials = (await session.execute(select(Material))).scalars().all()
        batches = (await session.execute(select(ElectrodeBatch))).scalars().all()
        edges = (await session.execute(select(LineageEdge))).scalars().all()
        builds = (await session.execute(select(CellBuild))).scalars().all()

    assert {material.name for material in materials} >= {"NMC811", "PVDF HSV1810", "Hx-T", "Hx-e"}
    assert {batch.batch_name for batch in batches} >= {"N1", "N1B", "N16a"}
    assert "N16" not in {batch.batch_name for batch in batches}
    assert any(edge.relationship_type == "branches_to" for edge in edges)
    assert any(build.build_name == "N1 i" for build in builds)
    assert any(build.build_name == "N16a i" for build in builds)
    n1_batch = next(batch for batch in batches if batch.batch_name == "N1")
    assert n1_batch.created_at.date().isoformat() == "2025-09-11"
    n1_build = next(build for build in builds if build.build_name == "N1 i")
    assert n1_build.build_date.isoformat() == "2025-09-12"
    assert n1_build.created_at.date().isoformat() == "2025-09-12"

    await engine.dispose()


async def test_normalize_legacy_formulation_preserves_sub_one_percent_components():
    normalized = _normalize_legacy_formulation(
        json.dumps(
            [
                {"Component": "NMC811 (MSE)", "Dry Mass Fraction (%)": 92.0},
                {"Component": "PVDF HSV1810", "Dry Mass Fraction (%)": 4.0},
                {"Component": "Super P", "Dry Mass Fraction (%)": 3.5},
                {"Component": "MWCNTs", "Dry Mass Fraction (%)": 0.5},
            ]
        )
    )

    mwcnt = next(component for component in normalized if component["Component"] == "MWCNTs")
    assert mwcnt["Dry Mass Fraction (%)"] == 0.5


async def test_preview_backfills_parent_loading_from_stable_legacy_child_cells(tmp_path):
    workbook_path = tmp_path / "NMC Cathode Formulations 2025.xlsx"
    legacy_db_path = tmp_path / "cellscope.db"
    build_workbook(workbook_path)
    build_legacy_db(legacy_db_path)
    append_n12_loading_backfill_case(workbook_path, legacy_db_path)

    engine, session_factory = create_session_factory(tmp_path)
    await initialize_backend_db(engine)

    async with session_factory() as session:
        preview = await import_nmc_cathode_workbook(
            session,
            workbook_path=str(workbook_path),
            legacy_db_path=str(legacy_db_path),
            persist=False,
        )

    n12_parent = next(item for item in preview["parent_batches"] if item["batch_name"] == "N12")
    assert n12_parent["metadata_json"]["safe_metrics"]["disc_mass_mg"] == 13.6
    assert n12_parent["metadata_json"]["loading_backfill"]["source"] == "legacy_cell_loadings"
    assert not any(
        item["record_name"] == "N12" and item["field"] == "disc_mass_mg"
        for item in preview["review_items"]
    )

    await engine.dispose()
