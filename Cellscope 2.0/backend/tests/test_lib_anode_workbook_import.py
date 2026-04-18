"""Tests for the LIB anode workbook ontology import service."""

from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.database import Base
from app.models.ontology import CellBuild, ElectrodeBatch, LineageEdge, Material
from app.services.lib_anode_workbook import import_lib_anode_workbook


pytestmark = pytest.mark.asyncio


def create_session_factory(tmp_path: Path):
    db_path = tmp_path / "ontology-import-test.db"
    engine = create_async_engine(f"sqlite+aiosqlite:///{db_path}")
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    return engine, session_factory


async def initialize_backend_db(engine) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    t21 = workbook.create_sheet("T21")
    t21["A1"] = "Experiment"
    t21["B1"] = 21
    t21["C1"] = "Notes"
    t21["D1"] = "Increase Si portion to balance capacity and stability."
    t21["A2"] = "Date"
    t21["B2"] = datetime(2025, 7, 16)
    t21["A3"] = "Component"
    t21["B3"] = "Dry Mass Fraction "
    t21["A4"] = "Graphite"
    t21["B4"] = 0.58
    t21["A5"] = "Cx (beta)"
    t21["B5"] = 0.07
    t21["A6"] = "Si"
    t21["B6"] = 0.22
    t21["A8"] = "LiPAA"
    t21["B8"] = 0.13
    t21["D15"] = "Active mass density, g/cc"
    t21["D16"] = 1.5
    t21["E15"] = "Slurry Density (g/mL)"
    t21["E16"] = 1.1
    t21["E18"] = "Disc Mass (mg)"
    t21["E19"] = 4.4
    t21["E20"] = "Electrode Density (g/mL)"
    t21["E21"] = 1.23
    t21["E22"] = "Porosity"
    t21["E23"] = 0.41
    t21["D24"] = "Pre-Press Thickness (um)"
    t21["E24"] = "Pressed Thickness (um)"
    t21["D25"] = 91
    t21["E25"] = 82
    t21["H34"] = "Loading"
    t21["I34"] = 10

    t8 = workbook.create_sheet("T8")
    t8["A1"] = "Experiment"
    t8["B1"] = "8a"
    t8["C1"] = "Notes"
    t8["D1"] = "Conventional Gr Cell Evaluation vs Super P"
    t8["E1"] = "Result Notes"
    t8["F1"] = "Electrodes successfully prepared."
    t8["H1"] = "Experiment"
    t8["I1"] = "8b"
    t8["J1"] = "Notes"
    t8["K1"] = "Conventional Gr Cell Evaluation vs Super P"
    t8["A2"] = "Date"
    t8["B2"] = datetime(2024, 11, 7)
    t8["A3"] = "Component"
    t8["B3"] = "Dry Mass Fraction "
    t8["H3"] = "Component"
    t8["I3"] = "Dry Mass Fraction "
    t8["A4"] = "Graphite"
    t8["B4"] = 0.9
    t8["A5"] = "Cx (beta)"
    t8["B5"] = 0.05
    t8["A7"] = "CMC"
    t8["B7"] = 0.0166
    t8["A8"] = "SBR"
    t8["B8"] = 0.0334
    t8["H4"] = "Graphite"
    t8["I4"] = 0.9
    t8["H5"] = "Super P "
    t8["I5"] = 0.05
    t8["H7"] = "CMC"
    t8["I7"] = 0.0166
    t8["H8"] = "SBR"
    t8["I8"] = 0.0334
    t8["E28"] = "Pre-press thickness (um)"
    t8["F28"] = "Post-Press Thickness (um)"
    t8["E29"] = 94
    t8["F29"] = 74
    t8["L28"] = "Pre-press thickness (um)"
    t8["M28"] = "Post-Press Thickness (um)"
    t8["L29"] = "--"
    t8["M29"] = 60

    workbook.create_sheet("Anode Experimental Data")
    workbook.create_sheet("Densities")
    workbook.save(path)


def build_legacy_db(path: Path) -> None:
    conn = sqlite3.connect(path)
    try:
        conn.executescript(
            """
            CREATE TABLE projects (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                project_type TEXT
            );

            CREATE TABLE cell_experiments (
                id INTEGER PRIMARY KEY,
                project_id INTEGER NOT NULL,
                cell_name TEXT NOT NULL,
                loading REAL,
                active_material REAL,
                electrolyte TEXT,
                formulation_json TEXT,
                solids_content REAL,
                pressed_thickness REAL,
                experiment_notes TEXT,
                substrate TEXT,
                separator TEXT,
                data_json TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO projects (id, name, project_type) VALUES (4, 'LIB Anodes', 'Anode')"
        )

        t21_data = {
            "experiment_date": "2025-07-16",
            "cells": [
                {
                    "cell_name": "T21 B2 i",
                    "test_number": "T21 B2 i",
                    "loading": 4.4,
                    "active_material": 80.0,
                    "electrolyte": "1.0M LiPF6",
                    "substrate": "Copper",
                    "separator": "Celgard",
                    "formulation": [
                        {"Component": "Graphite", "Dry Mass Fraction (%)": 58.0},
                        {"Component": "Cx (beta)", "Dry Mass Fraction (%)": 7.0},
                        {"Component": "Silicon", "Dry Mass Fraction (%)": 22.0},
                        {"Component": "LiPAA", "Dry Mass Fraction (%)": 13.0},
                    ],
                }
            ],
        }
        conn.execute(
            """
            INSERT INTO cell_experiments (
                id, project_id, cell_name, loading, active_material, electrolyte,
                formulation_json, solids_content, pressed_thickness, experiment_notes,
                substrate, separator, data_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                21,
                4,
                "T21 B2",
                None,
                None,
                "1.0M LiPF6",
                None,
                16.5,
                82.0,
                "Increase Si portion to balance capacity and stability. Uncal.",
                "Copper",
                "Celgard",
                json.dumps(t21_data),
            ),
        )
        conn.commit()
    finally:
        conn.close()


async def test_preview_import_creates_root_variants_and_legacy_children(tmp_path):
    workbook_path = tmp_path / "Anode Formulations 2024-25.xlsx"
    legacy_db_path = tmp_path / "cellscope.db"
    build_workbook(workbook_path)
    build_legacy_db(legacy_db_path)

    engine, session_factory = create_session_factory(tmp_path)
    await initialize_backend_db(engine)

    async with session_factory() as session:
        preview = await import_lib_anode_workbook(
            session,
            workbook_path=str(workbook_path),
            legacy_db_path=str(legacy_db_path),
            persist=False,
        )

    assert preview["persisted"] is False
    assert preview["counts"]["parent_batches"] == 1
    assert preview["counts"]["variant_batches"] == 3
    assert any(item["batch_name"] == "T8a" and item["parent_batch_name"] is None for item in preview["variant_batches"])
    assert any(item["batch_name"] == "T8b" and item["parent_batch_name"] is None for item in preview["variant_batches"])
    t21_variant = next(item for item in preview["variant_batches"] if item["batch_name"] == "T21 B2")
    assert t21_variant["parent_batch_name"] == "T21"
    t21_parent = next(item for item in preview["parent_batches"] if item["batch_name"] == "T21")
    assert t21_parent["metadata_json"]["study_family"] == "silicon_formulation_optimization"
    assert any(item["record_name"] == "T8b" and item["field"] == "pre_press_thickness_um" for item in preview["review_items"])
    assert any(item["name"] == "Graphite" for item in preview["materials"])
    assert any("multiple workbook sub-experiments" in note for note in preview["notes"])

    await engine.dispose()


async def test_persist_import_creates_anode_batches_lineage_and_cell_builds(tmp_path):
    workbook_path = tmp_path / "Anode Formulations 2024-25.xlsx"
    legacy_db_path = tmp_path / "cellscope.db"
    build_workbook(workbook_path)
    build_legacy_db(legacy_db_path)

    engine, session_factory = create_session_factory(tmp_path)
    await initialize_backend_db(engine)

    async with session_factory() as session:
        result = await import_lib_anode_workbook(
            session,
            workbook_path=str(workbook_path),
            legacy_db_path=str(legacy_db_path),
            persist=True,
        )
        await session.commit()

        batches = (await session.execute(select(ElectrodeBatch).order_by(ElectrodeBatch.batch_name))).scalars().all()
        materials = (await session.execute(select(Material).order_by(Material.name))).scalars().all()
        edges = (await session.execute(select(LineageEdge))).scalars().all()
        builds = (await session.execute(select(CellBuild).order_by(CellBuild.build_name))).scalars().all()

    assert result["persisted"] is True
    assert [batch.batch_name for batch in batches] == ["T21", "T21 B2", "T8a", "T8b"]
    assert all(batch.electrode_role.name == "ANODE" for batch in batches)
    assert any(material.name == "Graphite" for material in materials)
    assert any(edge.relationship_type == "branches_to" for edge in edges)
    assert any(edge.relationship_type == "built_into" for edge in edges)
    assert [build.build_name for build in builds] == ["T21 B2 i"]

    await engine.dispose()
