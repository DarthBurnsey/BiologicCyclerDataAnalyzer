"""Workbook-backed ontology import for the LIB anode formulation tracker."""

from __future__ import annotations

import csv
import json
import re
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from sqlalchemy import and_, delete, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ontology import (
    CellBuild,
    CellBuildStatus,
    ElectrodeBatch,
    ElectrodeRole,
    LineageEdge,
    LineageEntityType,
    Material,
    MaterialCategory,
    ProcessRun,
    ProcessType,
)

REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_LEGACY_DB_PATH = REPO_ROOT / "cellscope.db"

IGNORE_SHEETS = {
    "Anode Experimental Data",
    "Densities",
    "Viscosity Testing",
    "Brian",
    "Jong",
}
IMPORT_SOURCE = "lib_anode_workbook_import"
IMPORT_VERSION = "legacy-workbook-2026"
LEGACY_PROJECT_NAME = "LIB Anodes"
PRIMARY_REVIEW_FIELDS = {
    "disc_mass_mg",
    "loading_mg_cm2",
    "pre_press_thickness_um",
    "pressed_thickness_um",
}
METRIC_RULES: Dict[str, Tuple[str, str]] = {
    "active mass density, g/cc": ("active_mass_density_g_cc", "below"),
    "slurry density (g/ml)": ("slurry_density_g_ml", "below"),
    "electrode density (g/ml)": ("electrode_density_g_ml", "below"),
    "porosity": ("porosity", "below"),
    "disc mass (mg)": ("disc_mass_mg", "below"),
    "loading": ("loading_mg_cm2", "right"),
    "pre-press thickness (um)": ("pre_press_thickness_um", "below"),
    "post-press thickness (um)": ("pressed_thickness_um", "below"),
    "pressed thickness (um)": ("pressed_thickness_um", "below"),
    "press thickness reduction": ("thickness_reduction_pct", "below"),
    "mass of electrode": ("electrode_mass_mg", "right"),
}


def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_float(value: Any) -> Optional[float]:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value)
    if not text or text in {"--", ".", "-", "#N/A", "#DIV/0!", "#VALUE!", "#REF!"}:
        return None
    text = text.replace(",", "")
    text = text.rstrip("%")
    try:
        return float(text)
    except ValueError:
        return None


def _as_percent(value: Any) -> Optional[float]:
    number = _as_float(value)
    if number is None:
        return None
    return number * 100.0 if number <= 1.0 else number


def _as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    for candidate in (text[:10], text):
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00")).date()
        except ValueError:
            continue
    return None


def _as_datetime_from_date(value: Optional[date]) -> Optional[datetime]:
    if value is None:
        return None
    return datetime.combine(value, datetime.min.time())


def _normalize_label(value: Any) -> Optional[str]:
    text = _text(value)
    if not text:
        return None
    normalized = text.lower()
    normalized = normalized.replace("µ", "u")
    normalized = normalized.replace("%", "pct")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip(" .")


def _extract_numeric_candidates(value: Any) -> List[float]:
    number = _as_float(value)
    if number is not None:
        return [number]
    text = _text(value)
    if not text:
        return []
    cleaned = re.sub(r"-?\d+\s*mesh", "", text, flags=re.IGNORECASE)
    candidates = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", cleaned)]
    return candidates


def _normalize_metric_value(field: str, value: float) -> float:
    if field == "thickness_reduction_pct" and value <= 1.0:
        return value * 100.0
    return value


def _combine_notes(*values: Optional[str]) -> Optional[str]:
    parts: List[str] = []
    for value in values:
        text = _text(value)
        if text and text not in parts:
            parts.append(text)
    return "\n".join(parts) if parts else None


def _canonical_variant_key(name: str) -> Optional[str]:
    raw = _text(name)
    if not raw:
        return None
    raw = re.sub(r"\s*-\s*repaired\b", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s+", "", raw)
    return raw.upper()


def _display_variant_name(name: str) -> str:
    return re.sub(r"\s*-\s*repaired\b", "", name, flags=re.IGNORECASE).strip()


def _extract_parent_code(name: str) -> Optional[str]:
    raw = _text(name)
    if not raw:
        return None
    match = re.match(r"^(T\d+)", raw, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return _canonical_variant_key(raw)


def _infer_calendared(*values: Optional[str]) -> Optional[bool]:
    combined = " ".join(filter(None, (_text(value) for value in values))).lower()
    if not combined:
        return None
    if "uncal" in combined:
        return False
    if re.search(r"\bcal\b", combined):
        return True
    if "calender" in combined or "calendar" in combined:
        return True
    return None


def _flatten_issue_rows(issue_type: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for item in items:
        rows.append(
            {
                "issue_type": issue_type,
                "status": item.get("status"),
                "severity": item.get("severity"),
                "scope": item.get("scope"),
                "record_name": item.get("record_name"),
                "field": item.get("field"),
                "message": item.get("message"),
                "workbook_value": json.dumps(item.get("workbook_value"), default=str),
                "legacy_value": json.dumps(item.get("legacy_value"), default=str),
                "workbook_ref": item.get("workbook_ref"),
                "legacy_ref": item.get("legacy_ref"),
            }
        )
    return rows


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")


def _write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _write_markdown_summary(path: Path, plan: Dict[str, Any]) -> None:
    counts = plan["counts"]
    lines = [
        "# LIB Anode Ontology Reconciliation",
        "",
        f"- Source workbook: `{plan['source_workbook']}`",
        f"- Legacy database: `{plan['legacy_db_path']}`",
        "",
        "## Counts",
        "",
        f"- Parent batches: {counts['parent_batches']}",
        f"- Variant batches: {counts['variant_batches']}",
        f"- Materials: {counts['materials']}",
        f"- Review items: {counts['review_items']}",
        f"- Mismatch items: {counts['mismatch_items']}",
        "",
        "## What Needs Attention",
        "",
        "- Confirm workbook fields that are still blank, ambiguous, or split across multiple candidate values.",
        "- Review legacy-only experiment rows that do not map back to a workbook sheet.",
        "- Resolve real disagreements where workbook measurements and legacy experiment data both exist.",
        "",
        "## Files",
        "",
        "- `ontology_import_preview.json`: full preview payload",
        "- `ontology_import_issues.csv`: flattened review + mismatch issues",
        "- `ontology_import_parent_batches.json`: parent batch preview",
        "- `ontology_import_variant_batches.json`: variant batch preview",
        "- `ontology_import_cell_builds.json`: proposed cell-build lineage preview",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_review_export(plan: Dict[str, Any], export_dir: Path) -> Dict[str, str]:
    export_dir.mkdir(parents=True, exist_ok=True)

    preview_path = export_dir / "ontology_import_preview.json"
    issues_path = export_dir / "ontology_import_issues.csv"
    parent_batches_path = export_dir / "ontology_import_parent_batches.json"
    variant_batches_path = export_dir / "ontology_import_variant_batches.json"
    cell_builds_path = export_dir / "ontology_import_cell_builds.json"
    readme_path = export_dir / "README.md"

    _write_json(preview_path, plan)
    _write_json(parent_batches_path, plan["parent_batches"])
    _write_json(variant_batches_path, plan["variant_batches"])
    _write_json(cell_builds_path, plan["cell_builds"])
    _write_csv(
        issues_path,
        [
            *_flatten_issue_rows("review_required", plan["review_items"]),
            *_flatten_issue_rows("mismatch", plan["mismatch_items"]),
        ],
    )
    _write_markdown_summary(readme_path, plan)

    return {
        "export_dir": str(export_dir),
        "preview_json": str(preview_path),
        "issues_csv": str(issues_path),
        "parent_batches_json": str(parent_batches_path),
        "variant_batches_json": str(variant_batches_path),
        "cell_builds_json": str(cell_builds_path),
        "readme_md": str(readme_path),
    }


def _numbers_close(left: Optional[float], right: Optional[float], tolerance: float = 0.5) -> bool:
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def _create_issue(
    *,
    status: str,
    scope: str,
    record_name: str,
    field: str,
    message: str,
    workbook_value: Any = None,
    legacy_value: Any = None,
    workbook_ref: Optional[str] = None,
    legacy_ref: Optional[str] = None,
    severity: str = "warning",
) -> Dict[str, Any]:
    return {
        "status": status,
        "severity": severity,
        "scope": scope,
        "record_name": record_name,
        "field": field,
        "message": message,
        "workbook_value": workbook_value,
        "legacy_value": legacy_value,
        "workbook_ref": workbook_ref,
        "legacy_ref": legacy_ref,
    }


def _normalize_material_name(raw_name: str) -> Tuple[str, MaterialCategory, Dict[str, Any]]:
    text = _text(raw_name) or "Unknown"
    normalized = text.lower().replace("β", "beta")
    normalized = normalized.replace("-", " ")
    normalized = normalized.replace("(", " ").replace(")", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()

    metadata: Dict[str, Any] = {"source_name": text}
    if "btr graphite" in normalized:
        metadata["manufacturer_hint"] = "BTR"
        return "Graphite", MaterialCategory.ANODE_ACTIVE, metadata
    if "mse graphite" in normalized:
        metadata["manufacturer_hint"] = "MSE"
        return "Graphite", MaterialCategory.ANODE_ACTIVE, metadata
    if normalized == "graphite":
        return "Graphite", MaterialCategory.ANODE_ACTIVE, metadata
    if normalized in {"si", "silicon"} or normalized.startswith("si "):
        return "Silicon", MaterialCategory.ANODE_ACTIVE, metadata
    if "s03cd" in normalized or "si/c" in normalized:
        metadata["product_hint"] = text
        return "Si/C Composite", MaterialCategory.ANODE_ACTIVE, metadata
    if normalized.startswith("sicx200"):
        return "SiCx200", MaterialCategory.ANODE_ACTIVE, metadata
    if normalized.startswith("sicx100"):
        return "SiCx100", MaterialCategory.ANODE_ACTIVE, metadata
    if (
        normalized.startswith("sic150")
        or normalized.startswith("sic alpha")
        or normalized.startswith("sic beta")
        or normalized == "sic"
    ):
        metadata["product_hint"] = text
        return "SiC", MaterialCategory.ANODE_ACTIVE, metadata
    if "super p" in normalized:
        return "Super P", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "mwcnt" in normalized:
        return "MWCNTs", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "carbon nanotube" in normalized or normalized == "cnt":
        return "MWCNTs", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "cx beta" in normalized or "cx b" in normalized:
        metadata["manufacturer_hint"] = "Hexegen"
        return "Cx (beta)", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "cx echo" in normalized or "cx e" in normalized:
        metadata["manufacturer_hint"] = "Hexegen"
        return "Cx (echo)", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "hx e" in normalized or "hx echo" in normalized:
        metadata["manufacturer_hint"] = "Hexegen"
        if "aw" in normalized or "acid" in normalized or "hcl" in normalized:
            metadata["post_processing"] = "acid_wash"
        return "Hx-e", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if normalized == "cx":
        metadata["manufacturer_hint"] = "Hexegen"
        return "Cx", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "carboxymethyl cellulose" in normalized:
        return "CMC", MaterialCategory.BINDER, metadata
    if normalized == "cmc":
        return "CMC", MaterialCategory.BINDER, metadata
    if "styrene butadiene rubber" in normalized:
        return "SBR", MaterialCategory.BINDER, metadata
    if normalized == "sbr":
        return "SBR", MaterialCategory.BINDER, metadata
    if "lipaa" in normalized:
        return "LiPAA", MaterialCategory.BINDER, metadata
    if normalized.startswith("ast-9005 26"):
        return "AST-9005(26)", MaterialCategory.BINDER, metadata
    if normalized.startswith("ast-9005 30"):
        return "AST-9005(30)", MaterialCategory.BINDER, metadata
    if "b oh 3" in normalized:
        metadata["functional_role"] = "crosslinker"
        return "B(OH)3", MaterialCategory.OTHER, metadata
    if normalized in {"water", "h2o"}:
        return "Water", MaterialCategory.OTHER, metadata
    return text, MaterialCategory.OTHER, metadata


def _serialize_component(component: Dict[str, Any]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "Component": component["Component"],
        "Dry Mass Fraction (%)": round(component["Dry Mass Fraction (%)"], 4),
    }
    metadata = component.get("metadata") or {}
    if metadata:
        payload["metadata"] = metadata
    return payload


def _component_signature(components: Iterable[Dict[str, Any]]) -> List[Tuple[str, float]]:
    signature: List[Tuple[str, float]] = []
    for component in components:
        signature.append(
            (
                component["Component"],
                round(float(component["Dry Mass Fraction (%)"]), 2),
            )
        )
    return sorted(signature)


def _normalize_experiment_name(raw_name: Any, sheet_name: str) -> str:
    text = _text(raw_name) or sheet_name
    if re.match(r"^\d+[A-Za-z0-9]*$", text):
        prefix_match = re.match(r"^([A-Za-z]+)", sheet_name)
        prefix = prefix_match.group(1) if prefix_match else ""
        return f"{prefix}{text}"
    return text.replace("  ", " ").strip()


def _select_primary_active_material(formulation: Sequence[Dict[str, Any]]) -> Optional[str]:
    active_components = [
        component
        for component in formulation
        if component.get("category") == MaterialCategory.ANODE_ACTIVE
    ]
    if not active_components:
        return None
    primary = max(active_components, key=lambda item: item["Dry Mass Fraction (%)"])
    return primary["Component"]


def _sheet_number(name: str) -> Optional[int]:
    match = re.match(r"^T(\d+)", name, re.IGNORECASE)
    return int(match.group(1)) if match else None


def _family_for_batch(batch_name: str) -> str:
    number = _sheet_number(batch_name)
    if number is None:
        return "miscellaneous"
    if number <= 10:
        return "process_development"
    if number <= 21:
        return "silicon_formulation_optimization"
    if number <= 23:
        return "full_cell_and_adhesion_scaleup"
    if number <= 26:
        return "sicx_material_screening"
    if number <= 29:
        return "graphite_baseline_and_binder_screening"
    if number <= 32:
        return "sic_active_material_screening"
    if number <= 34:
        return "graphite_benchmarking_and_calendering"
    if number <= 39:
        return "jong_recipe_and_binder_screening"
    return "mixing_and_crack_mitigation"


def _infer_tags(text_blob: str, formulation: Sequence[Dict[str, Any]]) -> List[str]:
    lower = text_blob.lower()
    tags: List[str] = []
    keyword_tags = {
        "train": "training",
        "viscos": "viscosity",
        "solid": "solids_content",
        "adhesion": "adhesion",
        "delamination": "delamination",
        "full cell": "full_cell_candidate",
        "pouch": "pouch_cell_candidate",
        "sicx": "sicx",
        "sic ": "sic",
        "sic)": "sic",
        "graphite": "graphite",
        "btr": "btr_graphite",
        "electrolyte": "electrolyte_screen",
        "super p": "super_p_comparison",
        "binder": "binder_screen",
        "mix": "mixing_strategy",
        "crack": "cracking",
        "agglomer": "agglomeration_control",
        "cal": "calendering",
        "uncal": "uncalendered",
    }
    for needle, tag in keyword_tags.items():
        if needle in lower and tag not in tags:
            tags.append(tag)

    component_names = {component["Component"] for component in formulation}
    if "Silicon" in component_names:
        tags.append("silicon")
    if {"Graphite", "Silicon"} <= component_names:
        tags.append("silicon_graphite_blend")
    if "SiC" in component_names:
        tags.append("sic")
    if "SiCx100" in component_names or "SiCx200" in component_names:
        tags.append("sicx")
    if "Si/C Composite" in component_names:
        tags.append("si_c_composite")
    if "AST-9005(26)" in component_names or "AST-9005(30)" in component_names:
        tags.append("ast_binder")
    return sorted(set(tags))


def _infer_hypothesis(notes: Optional[str], result_notes: Optional[str]) -> Optional[str]:
    text_blob = " ".join(filter(None, (_text(notes), _text(result_notes)))).lower()
    if not text_blob:
        return None
    if "increase graphite" in text_blob:
        return "Higher graphite loading should improve mechanical stability and reduce failure."
    if "increase si" in text_blob:
        return "Higher silicon content should raise capacity while testing stability limits."
    if "super p" in text_blob and "comparison" in text_blob:
        return "The conductive additive choice should materially affect electrode quality and performance."
    if "improve adhesion" in text_blob or "delamination" in text_blob:
        return "Binder and dispersion changes should improve adhesion and reduce delamination."
    if "full cell" in text_blob or "pouch" in text_blob:
        return "This recipe is being pushed toward a stackable full-cell-compatible anode."
    if "btr graphite" in text_blob or "comparison against mse graphite" in text_blob:
        return "Alternative graphite sourcing and calendaring may change loading uniformity and downstream performance."
    if "jong" in text_blob or "ast-9005" in text_blob:
        return "Binder grade and agglomeration-control steps should determine whether the recipe coats cleanly without cracking."
    if "crack" in text_blob or "brittle" in text_blob:
        return "Binder-system and mixing changes should suppress brittle films and post-dry cracking."
    return None


def _derive_purpose_metadata(
    batch_name: str,
    *,
    notes: Optional[str],
    result_notes: Optional[str],
    formulation: Sequence[Dict[str, Any]],
    legacy_notes: Optional[str] = None,
    extra_tags: Optional[Sequence[str]] = None,
) -> Dict[str, Any]:
    purpose_summary = legacy_notes or notes
    family = _family_for_batch(batch_name)
    text_blob = " ".join(filter(None, (_text(notes), _text(result_notes), _text(legacy_notes))))
    study_tags = _infer_tags(text_blob, formulation)
    if extra_tags:
        study_tags.extend(extra_tags)
    metadata: Dict[str, Any] = {
        "study_family": family,
        "purpose_summary": _text(purpose_summary),
        "study_tags": sorted(set(study_tags)),
    }
    hypothesis = _infer_hypothesis(notes or legacy_notes, result_notes)
    if hypothesis:
        metadata["hypothesis"] = hypothesis
    if _text(result_notes):
        metadata["result_takeaway"] = _text(result_notes)
    return metadata


def _panel_descriptor(formulation: Sequence[Dict[str, Any]]) -> Optional[str]:
    components = {component["Component"] for component in formulation}
    if "Super P" in components and "Cx (beta)" not in components:
        return "Super P comparison arm"
    if "Cx (beta)" in components and "Super P" not in components:
        return "Cx (beta) comparison arm"
    return None


def _extract_panel_definitions(ws: Any, sheet_name: str) -> List[Dict[str, Any]]:
    panels: List[Dict[str, Any]] = []
    for col in range(1, ws.max_column + 1):
        if _normalize_label(ws.cell(1, col).value) != "experiment":
            continue
        batch_name = _normalize_experiment_name(ws.cell(1, col + 1).value, sheet_name)
        panel = {
            "batch_name": batch_name,
            "start_col": col,
            "notes": _text(ws.cell(1, col + 3).value),
            "result_notes": _text(ws.cell(1, col + 5).value),
            "build_date": _as_date(ws.cell(2, col + 1).value),
        }
        panels.append(panel)

    if not panels:
        panels.append(
            {
                "batch_name": sheet_name,
                "start_col": 1,
                "notes": _text(ws["D1"].value),
                "result_notes": _text(ws["F1"].value),
                "build_date": _as_date(ws["B2"].value),
            }
        )

    for index, panel in enumerate(panels):
        next_start = panels[index + 1]["start_col"] if index + 1 < len(panels) else ws.max_column + 1
        panel["end_col"] = next_start - 1
    return panels


def _parse_formulation(ws: Any, *, start_col: int) -> List[Dict[str, Any]]:
    header_row = None
    for row in range(1, min(8, ws.max_row) + 1):
        if _normalize_label(ws.cell(row, start_col).value) == "component":
            header_row = row
            break
    if header_row is None:
        return []

    formulation: List[Dict[str, Any]] = []
    for row in range(header_row + 1, min(header_row + 20, ws.max_row) + 1):
        raw_name = _text(ws.cell(row, start_col).value)
        if not raw_name:
            if formulation:
                break
            continue
        if _normalize_label(raw_name) == "total":
            break
        fraction_pct = _as_percent(ws.cell(row, start_col + 1).value)
        if fraction_pct is None or fraction_pct <= 0:
            continue
        canonical_name, category, metadata = _normalize_material_name(raw_name)
        formulation.append(
            {
                "Component": canonical_name,
                "Dry Mass Fraction (%)": fraction_pct,
                "category": category,
                "metadata": metadata,
            }
        )
    return formulation


def _collect_metric_candidates(
    value_ws: Any,
    *,
    col_start: int,
    col_end: int,
) -> Dict[str, List[Dict[str, Any]]]:
    hits: Dict[str, List[Dict[str, Any]]] = {}
    for row in range(1, value_ws.max_row + 1):
        for col in range(col_start, min(col_end, value_ws.max_column) + 1):
            label = _normalize_label(value_ws.cell(row, col).value)
            rule = METRIC_RULES.get(label or "")
            if not rule:
                continue
            field, direction = rule
            if direction == "below":
                value_row, value_col = row + 1, col
            else:
                value_row, value_col = row, col + 1
            if value_row > value_ws.max_row or value_col > value_ws.max_column:
                hits.setdefault(field, []).append(
                    {
                        "label_ref": f"{value_ws.title}!{value_ws.cell(row, col).coordinate}",
                        "value_ref": f"{value_ws.title}!{value_ws.cell(value_row, value_col).coordinate}",
                        "raw_value": None,
                        "values": [],
                    }
                )
                continue
            value_cell = value_ws.cell(value_row, value_col)
            candidates = [_normalize_metric_value(field, item) for item in _extract_numeric_candidates(value_cell.value)]
            hits.setdefault(field, []).append(
                {
                    "label_ref": f"{value_ws.title}!{value_ws.cell(row, col).coordinate}",
                    "value_ref": f"{value_ws.title}!{value_cell.coordinate}",
                    "raw_value": value_cell.value,
                    "values": candidates,
                }
            )
    return hits


def _resolve_sheet_measurements(
    value_ws: Any,
    *,
    record_name: str,
    col_start: int,
    col_end: int,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Dict[str, List[float]]]:
    metric_hits = _collect_metric_candidates(value_ws, col_start=col_start, col_end=col_end)
    safe_metrics: Dict[str, Any] = {}
    review_items: List[Dict[str, Any]] = []
    candidates_by_field: Dict[str, List[float]] = {}

    for field, hits in metric_hits.items():
        numeric_values = [
            value
            for hit in hits
            for value in hit["values"]
        ]
        unique_values = sorted({round(value, 4) for value in numeric_values})
        if unique_values:
            candidates_by_field[field] = unique_values

        if len(unique_values) == 1:
            safe_metrics[field] = unique_values[0]
            continue

        if len(unique_values) > 1:
            if field in PRIMARY_REVIEW_FIELDS:
                review_items.append(
                    _create_issue(
                        status="review_required",
                        scope="workbook_sheet",
                        record_name=record_name,
                        field=field,
                        message="Workbook contains multiple competing candidate values for this field.",
                        workbook_value=unique_values,
                        workbook_ref=", ".join(hit["value_ref"] for hit in hits),
                    )
                )
            continue

        if field in PRIMARY_REVIEW_FIELDS:
            first_hit = hits[0]
            review_items.append(
                _create_issue(
                    status="review_required",
                    scope="workbook_sheet",
                    record_name=record_name,
                    field=field,
                    message="Workbook field is blank, non-numeric, or otherwise unresolved.",
                    workbook_value=first_hit["raw_value"],
                    workbook_ref=first_hit["value_ref"],
                )
            )

    return safe_metrics, review_items, candidates_by_field


def _parse_workbook_batches(workbook_path: Path) -> Dict[str, Any]:
    workbook_values = load_workbook(workbook_path, data_only=True)

    sheet_records: Dict[str, Dict[str, Any]] = {}
    root_variants: List[Dict[str, Any]] = []
    materials: Dict[str, Dict[str, Any]] = {}
    review_items: List[Dict[str, Any]] = []
    notes: List[str] = []

    for sheet_name in workbook_values.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue
        ws_values = workbook_values[sheet_name]
        panels = _extract_panel_definitions(ws_values, sheet_name)

        sheet_notes = _combine_notes(*(panel["notes"] for panel in panels))
        sheet_result_notes = _combine_notes(*(panel["result_notes"] for panel in panels))
        sheet_date = next((panel["build_date"] for panel in panels if panel["build_date"]), _as_date(ws_values["B2"].value))
        process_name = f"{sheet_name} anode sheet process"

        if len(panels) > 1:
            notes.append(
                f"{sheet_name} contains multiple workbook sub-experiments ({', '.join(panel['batch_name'] for panel in panels)}); these are imported as root variant batches."
            )

        single_panel_maps_to_sheet = len(panels) == 1 and panels[0]["batch_name"].upper() == sheet_name.upper()
        parent_formulation: List[Dict[str, Any]] = []
        parent_metrics: Dict[str, Any] = {}
        parent_measurement_candidates: Dict[str, List[float]] = {}

        for panel in panels:
            formulation = _parse_formulation(ws_values, start_col=panel["start_col"])
            for component in formulation:
                materials[component["Component"]] = {
                    "name": component["Component"],
                    "category": component["category"],
                    "manufacturer": (component.get("metadata") or {}).get("manufacturer_hint"),
                    "description": None,
                    "metadata_json": component.get("metadata") or None,
                }

            safe_metrics, panel_review_items, measurement_candidates = _resolve_sheet_measurements(
                ws_values,
                record_name=panel["batch_name"],
                col_start=panel["start_col"],
                col_end=panel["end_col"],
            )
            review_items.extend(panel_review_items)

            descriptor = _panel_descriptor(formulation) if len(panels) > 1 else None
            purpose_metadata = _derive_purpose_metadata(
                panel["batch_name"],
                notes=panel["notes"] or sheet_notes,
                result_notes=panel["result_notes"] or sheet_result_notes,
                formulation=formulation,
                extra_tags=[descriptor] if descriptor else None,
            )

            if single_panel_maps_to_sheet:
                parent_formulation = formulation
                parent_metrics = safe_metrics
                parent_measurement_candidates = measurement_candidates
                sheet_records[sheet_name] = {
                    "batch_name": sheet_name,
                    "build_date": panel["build_date"] or sheet_date,
                    "notes": panel["notes"] or sheet_notes,
                    "result_notes": panel["result_notes"] or sheet_result_notes,
                    "formulation": formulation,
                    "safe_metrics": safe_metrics,
                    "measurement_candidates": measurement_candidates,
                    "source_sheet": sheet_name,
                    "process_name": process_name,
                    "purpose_metadata": purpose_metadata,
                }
                continue

            root_variants.append(
                {
                    "batch_name": panel["batch_name"],
                    "parent_batch_name": None,
                    "process_run_name": process_name,
                    "active_material_name": _select_primary_active_material(formulation),
                    "formulation_json": [_serialize_component(component) for component in formulation],
                    "notes": _combine_notes(panel["notes"], panel["result_notes"], descriptor),
                    "metadata_json": {
                        "source_workbook": str(workbook_path),
                        "source_sheet": sheet_name,
                        "source_kind": "workbook_panel",
                        "sheet_group_name": sheet_name,
                        "safe_metrics": safe_metrics,
                        "measurement_candidates": measurement_candidates or None,
                        "variant_descriptor": descriptor,
                        **purpose_metadata,
                    },
                    "source_kind": "workbook_panel_root",
                    "source_created_at": _as_datetime_from_date(panel["build_date"] or sheet_date),
                    "legacy_experiment_ids": [],
                }
            )

        if single_panel_maps_to_sheet and sheet_name not in sheet_records:
            sheet_records[sheet_name] = {
                "batch_name": sheet_name,
                "build_date": sheet_date,
                "notes": sheet_notes,
                "result_notes": sheet_result_notes,
                "formulation": parent_formulation,
                "safe_metrics": parent_metrics,
                "measurement_candidates": parent_measurement_candidates,
                "source_sheet": sheet_name,
                "process_name": process_name,
                "purpose_metadata": _derive_purpose_metadata(
                    sheet_name,
                    notes=sheet_notes,
                    result_notes=sheet_result_notes,
                    formulation=parent_formulation,
                ),
            }

    return {
        "sheet_records": sheet_records,
        "root_variants": root_variants,
        "materials": materials,
        "review_items": review_items,
        "notes": notes,
    }


def _normalize_legacy_formulation(raw_value: Any) -> List[Dict[str, Any]]:
    if raw_value is None or raw_value == "":
        return []
    data = raw_value
    if isinstance(raw_value, str):
        try:
            data = json.loads(raw_value)
        except json.JSONDecodeError:
            return []
    if not isinstance(data, list):
        return []

    parsed: List[Tuple[str, float]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        raw_name = _text(item.get("Component") or item.get("component"))
        raw_fraction = _as_float(
            item.get("Dry Mass Fraction (%)")
            or item.get("dry_mass_fraction_pct")
            or item.get("fraction")
        )
        if not raw_name or raw_fraction is None or raw_fraction <= 0:
            continue
        parsed.append((raw_name, raw_fraction))

    if not parsed:
        return []

    total = sum(value for _, value in parsed)
    scale = 100.0 if total <= 2.0 else 1.0
    normalized: List[Dict[str, Any]] = []
    for raw_name, raw_fraction in parsed:
        canonical_name, _, metadata = _normalize_material_name(raw_name)
        normalized.append(
            {
                "Component": canonical_name,
                "Dry Mass Fraction (%)": raw_fraction * scale,
                "metadata": metadata,
            }
        )
    return normalized


def _first_cell_formulation(data_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    cells = data_json.get("cells")
    if not isinstance(cells, list):
        return []
    for cell in cells:
        if not isinstance(cell, dict):
            continue
        normalized = _normalize_legacy_formulation(cell.get("formulation"))
        if normalized:
            return normalized
    return []


def _load_legacy_records(legacy_db_path: Path) -> Dict[str, Any]:
    conn = sqlite3.connect(legacy_db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT
                ce.id,
                ce.project_id,
                p.name AS project_name,
                p.project_type,
                ce.cell_name,
                ce.loading,
                ce.active_material,
                ce.electrolyte,
                ce.formulation_json,
                ce.solids_content,
                ce.pressed_thickness,
                ce.experiment_notes,
                ce.substrate,
                ce.separator,
                ce.data_json
            FROM cell_experiments ce
            JOIN projects p ON p.id = ce.project_id
            WHERE p.name = ?
            ORDER BY ce.id, ce.cell_name
            """,
            (LEGACY_PROJECT_NAME,),
        ).fetchall()
    finally:
        conn.close()

    rows_by_parent: Dict[str, List[Dict[str, Any]]] = {}
    all_records: List[Dict[str, Any]] = []
    for row in rows:
        data_json: Dict[str, Any] = {}
        try:
            if row["data_json"]:
                parsed = json.loads(row["data_json"])
                if isinstance(parsed, dict):
                    data_json = parsed
        except json.JSONDecodeError:
            data_json = {}

        formulation = _normalize_legacy_formulation(row["formulation_json"]) or _first_cell_formulation(data_json)
        tracking = data_json.get("tracking") if isinstance(data_json.get("tracking"), dict) else {}
        notes = _combine_notes(_text(row["experiment_notes"]), _text(tracking.get("notes")))
        record = {
            "id": row["id"],
            "project_id": row["project_id"],
            "project_name": row["project_name"],
            "project_type": row["project_type"],
            "cell_name": row["cell_name"],
            "match_key": _canonical_variant_key(row["cell_name"]),
            "parent_code": _extract_parent_code(row["cell_name"] or ""),
            "loading": _as_float(row["loading"]),
            "active_material_pct": _as_percent(row["active_material"]),
            "electrolyte": _text(row["electrolyte"]),
            "formulation": formulation,
            "solids_content_pct": _as_percent(row["solids_content"]),
            "pressed_thickness_um": _as_float(row["pressed_thickness"]),
            "notes": notes,
            "substrate": _text(row["substrate"]),
            "separator": _text(row["separator"]),
            "data_json": data_json,
            "calendared": _infer_calendared(notes),
            "is_repaired": "repaired" in (_text(row["cell_name"]) or "").lower(),
        }
        rows_by_parent.setdefault(record["parent_code"] or record["match_key"], []).append(record)
        all_records.append(record)

    return {
        "rows_by_parent": rows_by_parent,
        "records": all_records,
    }


def _collect_legacy_loading_values(records: Sequence[Dict[str, Any]]) -> List[float]:
    values: List[float] = []
    for record in records:
        cells = record.get("data_json", {}).get("cells")
        if isinstance(cells, list):
            for cell in cells:
                if not isinstance(cell, dict):
                    continue
                value = _as_float(cell.get("loading"))
                if value is not None:
                    values.append(value)
        top_level = _as_float(record.get("loading"))
        if top_level is not None and not cells:
            values.append(top_level)
    return values


def _derive_stable_legacy_loading(
    records: Sequence[Dict[str, Any]],
    *,
    tolerance: float = 0.35,
) -> Optional[Dict[str, Any]]:
    values = _collect_legacy_loading_values(records)
    if not values:
        return None
    if max(values) - min(values) > tolerance:
        return None
    representative = round(sum(values) / len(values), 4)
    return {
        "value": representative,
        "sample_count": len(values),
        "record_ids": [record["id"] for record in records],
        "source": "legacy_cell_loadings",
    }


def _choose_primary_legacy_record(records: Sequence[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not records:
        return None
    repaired = [record for record in records if record["is_repaired"]]
    candidates = repaired or list(records)
    return sorted(candidates, key=lambda record: (record["id"], record["cell_name"]))[-1]


def _merge_json(existing: Optional[Dict[str, Any]], incoming: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not existing and not incoming:
        return None
    payload = dict(existing or {})
    payload.update(incoming or {})
    return payload


def _compare_formulations(
    *,
    record_name: str,
    scope: str,
    workbook_formulation: Sequence[Dict[str, Any]],
    legacy_record: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    # Legacy LIB anode formulation payloads are incomplete and inconsistent.
    # The workbook is the authoritative source for formulation identity, so we
    # do not emit formulation mismatches from the legacy database.
    return []


def _create_cell_build_specs(
    *,
    source_batch_name: str,
    legacy_record: Dict[str, Any],
    relationship_type: str,
    relationship_notes: Optional[str] = None,
) -> List[Dict[str, Any]]:
    cells = legacy_record.get("data_json", {}).get("cells")
    if not isinstance(cells, list) or not cells:
        return []

    tracking = legacy_record.get("data_json", {}).get("tracking")
    tracking_date = None
    if isinstance(tracking, dict):
        tracking_date = _as_date(tracking.get("tracking_date"))
    build_date = _as_date(legacy_record.get("data_json", {}).get("experiment_date")) or tracking_date

    specs: List[Dict[str, Any]] = []
    for index, cell in enumerate(cells, start=1):
        if not isinstance(cell, dict):
            continue
        build_name = (
            _text(cell.get("cell_name"))
            or _text(cell.get("test_number"))
            or f"{legacy_record['cell_name']}::{index}"
        )
        specs.append(
            {
                "build_name": build_name,
                "chemistry": "LIB anode half-cell",
                "form_factor": None,
                "build_date": build_date,
                "status": CellBuildStatus.TESTING,
                "notes": _combine_notes(legacy_record.get("notes"), relationship_notes),
                "legacy_project_id": legacy_record["project_id"],
                "legacy_experiment_id": legacy_record["id"],
                "legacy_cell_id": None,
                "legacy_test_number": _text(cell.get("test_number")),
                "metadata_json": {
                    "legacy_project_name": legacy_record["project_name"],
                    "legacy_cell_name": legacy_record["cell_name"],
                    "legacy_test_number": _text(cell.get("test_number")),
                    "electrolyte": _text(cell.get("electrolyte")) or legacy_record.get("electrolyte"),
                    "separator": _text(cell.get("separator")) or legacy_record.get("separator"),
                    "substrate": _text(cell.get("substrate")) or legacy_record.get("substrate"),
                    "loading": _as_float(cell.get("loading")) or legacy_record.get("loading"),
                    "active_material_pct": _as_percent(cell.get("active_material")) or legacy_record.get("active_material_pct"),
                    "source_batch_name": source_batch_name,
                    "tracking_date": tracking_date.isoformat() if tracking_date else None,
                },
                "source_batch_name": source_batch_name,
            }
        )
    return specs


def _variant_descriptor_from_legacy(name: str, notes: Optional[str]) -> Optional[str]:
    combined = " ".join(filter(None, (_text(name), _text(notes)))).lower()
    if "repaired" in combined:
        return "repaired remake"
    if "uncal" in combined:
        return "uncalendered arm"
    if re.search(r"\bcal\b", combined) or "calender" in combined:
        return "calendered arm"
    if "electrolyte" in combined or "different el" in combined:
        return "electrolyte screen arm"
    return None


def _prepare_import_plan(
    workbook_path: Path,
    legacy_db_path: Path,
) -> Dict[str, Any]:
    workbook = _parse_workbook_batches(workbook_path)
    legacy = _load_legacy_records(legacy_db_path)

    materials = dict(workbook["materials"])
    process_runs: List[Dict[str, Any]] = []
    parent_batches: List[Dict[str, Any]] = []
    variant_batches: List[Dict[str, Any]] = list(workbook["root_variants"])
    experiment_links: List[Dict[str, Any]] = []
    cell_builds: List[Dict[str, Any]] = []
    review_items: List[Dict[str, Any]] = list(workbook["review_items"])
    mismatch_items: List[Dict[str, Any]] = []
    notes: List[str] = list(workbook["notes"])

    sheet_records = workbook["sheet_records"]
    seen_legacy_keys: set[str] = set()

    for sheet_name, sheet in sorted(sheet_records.items()):
        exact_legacy_rows = [
            record
            for record in legacy["rows_by_parent"].get(sheet_name, [])
            if record["match_key"] == sheet_name.upper()
        ]
        exact_legacy = _choose_primary_legacy_record(exact_legacy_rows)
        legacy_children_by_key: Dict[str, List[Dict[str, Any]]] = {}
        for record in legacy["rows_by_parent"].get(sheet_name, []):
            seen_legacy_keys.add(record["match_key"])
            if record["match_key"] == sheet_name.upper():
                continue
            legacy_children_by_key.setdefault(record["match_key"], []).append(record)

        process_runs.append(
            {
                "name": sheet["process_name"],
                "process_type": ProcessType.COATING,
                "started_at": sheet["build_date"],
                "completed_at": None,
                "settings_json": {
                    "source_workbook": str(workbook_path),
                    "source_sheet": sheet["source_sheet"],
                    "safe_metrics": sheet["safe_metrics"],
                    "measurement_candidates": sheet["measurement_candidates"] or None,
                    "study_family": sheet["purpose_metadata"].get("study_family"),
                },
                "notes": _combine_notes(sheet["notes"], sheet["result_notes"]),
            }
        )

        parent_notes = _combine_notes(sheet["notes"], sheet["result_notes"], exact_legacy.get("notes") if exact_legacy else None)
        parent_metadata = {
            "source_workbook": str(workbook_path),
            "source_sheet": sheet["source_sheet"],
            "safe_metrics": sheet["safe_metrics"],
            "measurement_candidates": sheet["measurement_candidates"] or None,
            "legacy_aliases": [record["cell_name"] for record in exact_legacy_rows],
            **sheet["purpose_metadata"],
        }
        if exact_legacy:
            parent_metadata["calendared"] = exact_legacy.get("calendared")
            parent_metadata["separator"] = exact_legacy.get("separator")
            parent_metadata["substrate"] = exact_legacy.get("substrate")

        active_material_name = _select_primary_active_material(sheet["formulation"])
        if active_material_name:
            parent_metadata["primary_active_material"] = active_material_name

        parent_batches.append(
            {
                "batch_name": sheet_name,
                "parent_batch_name": None,
                "process_run_name": sheet["process_name"],
                "active_material_name": active_material_name,
                "formulation_json": [_serialize_component(component) for component in sheet["formulation"]],
                "notes": parent_notes,
                "metadata_json": parent_metadata,
                "source_kind": "workbook_parent",
                "source_created_at": _as_datetime_from_date(sheet["build_date"]),
                "legacy_experiment_ids": [record["id"] for record in exact_legacy_rows],
            }
        )

        if exact_legacy:
            experiment_links.append(
                {
                    "source_batch_name": sheet_name,
                    "legacy_experiment_id": exact_legacy["id"],
                    "relationship_type": "evaluated_in",
                    "notes": "Workbook experiment sheet matched an exact legacy LIB Anodes row.",
                }
            )
            cell_builds.extend(
                _create_cell_build_specs(
                    source_batch_name=sheet_name,
                    legacy_record=exact_legacy,
                    relationship_type="built_into",
                )
            )
            mismatch_items.extend(
                _compare_formulations(
                    record_name=sheet_name,
                    scope="parent_batch",
                    workbook_formulation=parent_batches[-1]["formulation_json"],
                    legacy_record=exact_legacy,
                )
            )

            stable_exact_loading = _derive_stable_legacy_loading(exact_legacy_rows)
            workbook_disc_mass = sheet["safe_metrics"].get("disc_mass_mg")
            if (
                stable_exact_loading
                and workbook_disc_mass is not None
                and not _numbers_close(workbook_disc_mass, stable_exact_loading["value"])
            ):
                mismatch_items.append(
                    _create_issue(
                        status="mismatch",
                        scope="parent_batch",
                        record_name=sheet_name,
                        field="disc_mass_mg",
                        message="Workbook disc mass does not match stable legacy loading data.",
                        workbook_value=workbook_disc_mass,
                        legacy_value=stable_exact_loading["value"],
                        legacy_ref=f"cell_experiments:{exact_legacy['id']}",
                    )
                )
            workbook_pressed = sheet["safe_metrics"].get("pressed_thickness_um")
            if (
                workbook_pressed is not None
                and exact_legacy.get("pressed_thickness_um") is not None
                and not _numbers_close(workbook_pressed, exact_legacy["pressed_thickness_um"], tolerance=1.0)
            ):
                mismatch_items.append(
                    _create_issue(
                        status="mismatch",
                        scope="parent_batch",
                        record_name=sheet_name,
                        field="pressed_thickness_um",
                        message="Workbook pressed thickness does not match the legacy experiment row.",
                        workbook_value=workbook_pressed,
                        legacy_value=exact_legacy["pressed_thickness_um"],
                        legacy_ref=f"cell_experiments:{exact_legacy['id']}",
                    )
                )

        elif not legacy["rows_by_parent"].get(sheet_name):
            notes.append(
                f"{sheet_name} is workbook-only for now; the importer persists it as a canonical anode batch without a legacy experiment link."
            )

        for variant_key, related_legacy in sorted(legacy_children_by_key.items()):
            primary_legacy = _choose_primary_legacy_record(related_legacy)
            if primary_legacy is None:
                continue
            stable_loading = _derive_stable_legacy_loading(related_legacy)
            variant_name = _display_variant_name(primary_legacy["cell_name"])
            descriptor = _variant_descriptor_from_legacy(variant_name, primary_legacy.get("notes"))
            variant_formulation = primary_legacy["formulation"] or [
                {
                    "Component": component["Component"],
                    "Dry Mass Fraction (%)": component["Dry Mass Fraction (%)"],
                    "metadata": component.get("metadata") or {},
                }
                for component in parent_batches[-1]["formulation_json"]
            ]
            purpose_metadata = _derive_purpose_metadata(
                variant_name,
                notes=sheet["notes"],
                result_notes=sheet["result_notes"],
                formulation=variant_formulation,
                legacy_notes=primary_legacy.get("notes"),
                extra_tags=[descriptor] if descriptor else None,
            )

            for component in variant_formulation:
                canonical_name, category, metadata = _normalize_material_name(component["Component"])
                materials[canonical_name] = {
                    "name": canonical_name,
                    "category": category,
                    "manufacturer": (metadata or {}).get("manufacturer_hint"),
                    "description": None,
                    "metadata_json": metadata or None,
                }

            variant_metadata = {
                "source_workbook": str(workbook_path),
                "source_sheet": sheet["source_sheet"],
                "parent_batch_name": sheet_name,
                "legacy_aliases": [record["cell_name"] for record in related_legacy],
                "loading_backfill": stable_loading,
                "electrolyte": primary_legacy.get("electrolyte"),
                "separator": primary_legacy.get("separator"),
                "substrate": primary_legacy.get("substrate"),
                "calendared": primary_legacy.get("calendared"),
                "variant_descriptor": descriptor,
                **purpose_metadata,
            }
            if stable_loading:
                variant_metadata["loading_mg"] = stable_loading["value"]

            variant_batches.append(
                {
                    "batch_name": variant_name,
                    "parent_batch_name": sheet_name,
                    "process_run_name": sheet["process_name"],
                    "active_material_name": _select_primary_active_material(variant_formulation) or active_material_name,
                    "formulation_json": [_serialize_component(component) for component in variant_formulation],
                    "notes": _combine_notes(primary_legacy.get("notes"), descriptor),
                    "metadata_json": variant_metadata,
                    "source_kind": "legacy_child_variant",
                    "source_created_at": _as_datetime_from_date(sheet["build_date"]),
                    "legacy_experiment_ids": [record["id"] for record in related_legacy],
                }
            )

            experiment_links.append(
                {
                    "source_batch_name": variant_name,
                    "legacy_experiment_id": primary_legacy["id"],
                    "relationship_type": "evaluated_in",
                    "notes": "Legacy LIB Anodes experiment grouped under the workbook parent batch.",
                }
            )
            cell_builds.extend(
                _create_cell_build_specs(
                    source_batch_name=variant_name,
                    legacy_record=primary_legacy,
                    relationship_type="built_into",
                )
            )
            mismatch_items.extend(
                _compare_formulations(
                    record_name=variant_name,
                    scope="variant_batch",
                    workbook_formulation=variant_batches[-1]["formulation_json"],
                    legacy_record=primary_legacy,
                )
            )

    for root_variant in variant_batches:
        for component in root_variant.get("formulation_json") or []:
            canonical_name, category, metadata = _normalize_material_name(component["Component"])
            materials[canonical_name] = {
                "name": canonical_name,
                "category": category,
                "manufacturer": (metadata or {}).get("manufacturer_hint"),
                "description": None,
                "metadata_json": metadata or None,
            }

    for record in legacy["records"]:
        if record["match_key"] in seen_legacy_keys:
            continue
        mismatch_items.append(
            _create_issue(
                status="legacy_only",
                scope="legacy_experiment",
                record_name=record["cell_name"],
                field="workbook_match",
                message="Legacy LIB Anodes experiment has no matching workbook sheet.",
                legacy_ref=f"cell_experiments:{record['id']}",
            )
        )

    counts = {
        "materials": len(materials),
        "protocol_versions": 0,
        "process_runs": len(process_runs),
        "parent_batches": len(parent_batches),
        "variant_batches": len(variant_batches),
        "legacy_experiment_links": len(experiment_links),
        "cell_builds": len(cell_builds),
        "review_items": len(review_items),
        "mismatch_items": len(mismatch_items),
    }

    return {
        "source_workbook": str(workbook_path),
        "legacy_db_path": str(legacy_db_path),
        "counts": counts,
        "materials": sorted(materials.values(), key=lambda item: item["name"]),
        "protocol_versions": [],
        "process_runs": process_runs,
        "parent_batches": sorted(parent_batches, key=lambda item: item["batch_name"]),
        "variant_batches": sorted(variant_batches, key=lambda item: item["batch_name"]),
        "legacy_experiment_links": experiment_links,
        "cell_builds": sorted(cell_builds, key=lambda item: item["build_name"]),
        "review_items": review_items,
        "mismatch_items": mismatch_items,
        "notes": notes,
    }


async def _get_by_unique_field(
    db: AsyncSession,
    model: Any,
    field_name: str,
    value: Any,
) -> Optional[Any]:
    result = await db.execute(select(model).where(getattr(model, field_name) == value))
    return result.scalars().first()


async def _upsert_material(db: AsyncSession, spec: Dict[str, Any]) -> Material:
    material = await _get_by_unique_field(db, Material, "name", spec["name"])
    if material is None:
        material = Material(**spec)
        db.add(material)
        await db.flush()
        return material

    material.category = spec["category"]
    material.manufacturer = spec.get("manufacturer") or material.manufacturer
    material.description = spec.get("description") or material.description
    material.metadata_json = _merge_json(material.metadata_json, spec.get("metadata_json"))
    await db.flush()
    return material


async def _upsert_process_run(db: AsyncSession, spec: Dict[str, Any]) -> ProcessRun:
    process_run = await _get_by_unique_field(db, ProcessRun, "name", spec["name"])
    payload = {
        "name": spec["name"],
        "process_type": spec["process_type"],
        "protocol_version_id": None,
        "started_at": spec.get("started_at"),
        "completed_at": spec.get("completed_at"),
        "settings_json": spec.get("settings_json"),
        "notes": spec.get("notes"),
    }
    if process_run is None:
        process_run = ProcessRun(**payload)
        db.add(process_run)
        await db.flush()
        return process_run

    process_run.process_type = spec["process_type"]
    process_run.started_at = spec.get("started_at") or process_run.started_at
    process_run.completed_at = spec.get("completed_at") or process_run.completed_at
    process_run.settings_json = _merge_json(process_run.settings_json, spec.get("settings_json"))
    process_run.notes = spec.get("notes") or process_run.notes
    await db.flush()
    return process_run


async def _upsert_electrode_batch(
    db: AsyncSession,
    spec: Dict[str, Any],
    *,
    process_run_id: Optional[int],
    active_material_id: Optional[int],
) -> ElectrodeBatch:
    batch = await _get_by_unique_field(db, ElectrodeBatch, "batch_name", spec["batch_name"])
    payload = {
        "batch_name": spec["batch_name"],
        "electrode_role": ElectrodeRole.ANODE,
        "active_material_id": active_material_id,
        "process_run_id": process_run_id,
        "formulation_json": spec.get("formulation_json"),
        "notes": spec.get("notes"),
        "metadata_json": spec.get("metadata_json"),
        "created_at": spec.get("source_created_at"),
    }
    if batch is None:
        batch = ElectrodeBatch(**payload)
        db.add(batch)
        await db.flush()
        return batch

    batch.active_material_id = active_material_id
    batch.process_run_id = process_run_id
    batch.formulation_json = spec.get("formulation_json") or batch.formulation_json
    batch.notes = spec.get("notes") or batch.notes
    batch.metadata_json = _merge_json(batch.metadata_json, spec.get("metadata_json"))
    if spec.get("source_created_at") is not None:
        batch.created_at = spec["source_created_at"]
    await db.flush()
    return batch


async def _upsert_cell_build(db: AsyncSession, spec: Dict[str, Any]) -> CellBuild:
    source_created_at = _as_datetime_from_date(spec.get("build_date"))
    build = await _get_by_unique_field(db, CellBuild, "build_name", spec["build_name"])
    payload = {
        "build_name": spec["build_name"],
        "chemistry": spec.get("chemistry"),
        "form_factor": spec.get("form_factor"),
        "build_date": spec.get("build_date"),
        "status": spec.get("status", CellBuildStatus.PLANNED),
        "notes": spec.get("notes"),
        "legacy_project_id": spec.get("legacy_project_id"),
        "legacy_experiment_id": spec.get("legacy_experiment_id"),
        "legacy_cell_id": spec.get("legacy_cell_id"),
        "legacy_test_number": spec.get("legacy_test_number"),
        "metadata_json": spec.get("metadata_json"),
        "created_at": source_created_at,
    }
    if build is None:
        build = CellBuild(**payload)
        db.add(build)
        await db.flush()
        return build

    build.chemistry = spec.get("chemistry") or build.chemistry
    build.form_factor = spec.get("form_factor") or build.form_factor
    build.build_date = spec.get("build_date") or build.build_date
    build.status = spec.get("status", build.status)
    build.notes = spec.get("notes") or build.notes
    build.legacy_project_id = spec.get("legacy_project_id") or build.legacy_project_id
    build.legacy_experiment_id = spec.get("legacy_experiment_id") or build.legacy_experiment_id
    build.legacy_cell_id = spec.get("legacy_cell_id") or build.legacy_cell_id
    build.legacy_test_number = spec.get("legacy_test_number") or build.legacy_test_number
    build.metadata_json = _merge_json(build.metadata_json, spec.get("metadata_json"))
    if source_created_at is not None:
        build.created_at = source_created_at
    await db.flush()
    return build


async def _upsert_lineage_edge(
    db: AsyncSession,
    *,
    parent_type: LineageEntityType,
    parent_id: int,
    child_type: LineageEntityType,
    child_id: int,
    relationship_type: str,
    source: str,
    confidence: Optional[float] = None,
    notes: Optional[str] = None,
) -> LineageEdge:
    result = await db.execute(
        select(LineageEdge)
        .where(LineageEdge.parent_type == parent_type)
        .where(LineageEdge.parent_id == parent_id)
        .where(LineageEdge.child_type == child_type)
        .where(LineageEdge.child_id == child_id)
        .where(LineageEdge.relationship_type == relationship_type)
    )
    edge = result.scalars().first()
    if edge is None:
        edge = LineageEdge(
            parent_type=parent_type,
            parent_id=parent_id,
            child_type=child_type,
            child_id=child_id,
            relationship_type=relationship_type,
            source=source,
            confidence=confidence,
            notes=notes,
        )
        db.add(edge)
        await db.flush()
        return edge

    edge.source = source
    edge.confidence = confidence if confidence is not None else edge.confidence
    edge.notes = notes or edge.notes
    await db.flush()
    return edge


async def _remove_batch_and_edges(db: AsyncSession, batch_name: str) -> None:
    batch = await _get_by_unique_field(db, ElectrodeBatch, "batch_name", batch_name)
    if batch is None:
        return
    await db.execute(
        delete(LineageEdge).where(
            or_(
                and_(
                    LineageEdge.parent_type == LineageEntityType.ELECTRODE_BATCH,
                    LineageEdge.parent_id == batch.id,
                ),
                and_(
                    LineageEdge.child_type == LineageEntityType.ELECTRODE_BATCH,
                    LineageEdge.child_id == batch.id,
                ),
            )
        )
    )
    await db.delete(batch)
    await db.flush()


async def import_lib_anode_workbook(
    db: AsyncSession,
    *,
    workbook_path: str,
    legacy_db_path: Optional[str] = None,
    export_dir: Optional[str] = None,
    persist: bool = False,
) -> Dict[str, Any]:
    workbook = Path(workbook_path).expanduser().resolve()
    legacy_db = Path(legacy_db_path).expanduser().resolve() if legacy_db_path else DEFAULT_LEGACY_DB_PATH.resolve()
    review_export_dir = Path(export_dir).expanduser().resolve() if export_dir else None

    if not workbook.exists():
        raise ValueError(f"Workbook not found: {workbook}")
    if not legacy_db.exists():
        raise ValueError(f"Legacy database not found: {legacy_db}")

    plan = _prepare_import_plan(workbook, legacy_db)
    plan["persisted"] = False
    if review_export_dir is not None:
        plan["review_export"] = _write_review_export(plan, review_export_dir)
    if not persist:
        return plan

    material_map: Dict[str, Material] = {}
    for material_spec in plan["materials"]:
        material = await _upsert_material(db, material_spec)
        material_map[material.name] = material

    process_map: Dict[str, ProcessRun] = {}
    for process_spec in plan["process_runs"]:
        process_run = await _upsert_process_run(db, process_spec)
        process_map[process_run.name] = process_run

    batch_map: Dict[str, ElectrodeBatch] = {}
    for batch_spec in [*plan["parent_batches"], *plan["variant_batches"]]:
        process_run = process_map.get(batch_spec["process_run_name"])
        active_material = material_map.get(batch_spec.get("active_material_name"))
        batch = await _upsert_electrode_batch(
            db,
            batch_spec,
            process_run_id=process_run.id if process_run else None,
            active_material_id=active_material.id if active_material else None,
        )
        batch_map[batch.batch_name] = batch

    for batch_spec in [*plan["parent_batches"], *plan["variant_batches"]]:
        batch = batch_map[batch_spec["batch_name"]]
        parent_batch_name = batch_spec.get("parent_batch_name")
        if parent_batch_name:
            parent_batch = batch_map.get(parent_batch_name)
            if parent_batch:
                await _upsert_lineage_edge(
                    db,
                    parent_type=LineageEntityType.ELECTRODE_BATCH,
                    parent_id=parent_batch.id,
                    child_type=LineageEntityType.ELECTRODE_BATCH,
                    child_id=batch.id,
                    relationship_type="branches_to",
                    source=IMPORT_SOURCE,
                    confidence=1.0,
                    notes="Workbook parent/variant relationship.",
                )
        for component in batch_spec.get("formulation_json") or []:
            material = material_map.get(component.get("Component"))
            if not material:
                continue
            await _upsert_lineage_edge(
                db,
                parent_type=LineageEntityType.MATERIAL,
                parent_id=material.id,
                child_type=LineageEntityType.ELECTRODE_BATCH,
                child_id=batch.id,
                relationship_type="formulates_into",
                source=IMPORT_SOURCE,
                confidence=1.0,
                notes=f"Imported formulation component for {batch.batch_name}.",
            )

    for link_spec in plan["legacy_experiment_links"]:
        batch = batch_map.get(link_spec["source_batch_name"])
        if batch is None:
            continue
        await _upsert_lineage_edge(
            db,
            parent_type=LineageEntityType.ELECTRODE_BATCH,
            parent_id=batch.id,
            child_type=LineageEntityType.EXPERIMENT,
            child_id=link_spec["legacy_experiment_id"],
            relationship_type=link_spec["relationship_type"],
            source=IMPORT_SOURCE,
            confidence=0.9,
            notes=link_spec.get("notes"),
        )

    for build_spec in plan["cell_builds"]:
        batch = batch_map.get(build_spec["source_batch_name"])
        if batch is None:
            continue
        build = await _upsert_cell_build(db, build_spec)
        await _upsert_lineage_edge(
            db,
            parent_type=LineageEntityType.ELECTRODE_BATCH,
            parent_id=batch.id,
            child_type=LineageEntityType.CELL_BUILD,
            child_id=build.id,
            relationship_type="built_into",
            source=IMPORT_SOURCE,
            confidence=0.9,
            notes="Imported LIB anode cell-build lineage.",
        )

    plan["persisted"] = True
    return plan
