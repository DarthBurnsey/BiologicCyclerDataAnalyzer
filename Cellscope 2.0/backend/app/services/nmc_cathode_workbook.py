"""Workbook-backed ontology import for the NMC cathode formulation tracker."""

from __future__ import annotations

import copy
import csv
import json
import re
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import and_, delete, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ontology import (
    CellBuild,
    CellBuildStatus,
    ElectrodeBatch,
    ElectrodeRole,
    LineageEdge,
    LineageEntityType,
    Material,
    MaterialCategory,
    ProcessRun,
    ProcessType,
    ProtocolType,
    ProtocolVersion,
)

REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_LEGACY_DB_PATH = REPO_ROOT / "cellscope.db"

SUMMARY_SHEET = "Experimental Data"
IGNORE_SHEETS = {SUMMARY_SHEET, "Density"}
REVIEW_FILL_COLORS = {"FFFFFF00", "FFF4B084"}
IMPORT_SOURCE = "nmc_cathode_workbook_import"
IMPORT_VERSION = "legacy-workbook-2025"
FULL_CELL_OVERRIDES = {
    "FC4A": {
        "source_batch_name": "N12",
        "calendared": True,
        "notes": "Manual mapping from user: FC4a uses N12 as the calendared cathode.",
    },
    "FC5": {
        "source_batch_name": "N10",
        "calendared": True,
        "notes": (
            "Manual mapping from user: FC5 uses N10 as the calendared cathode; "
            "T35 is the calendared anode."
        ),
    },
}
EXPECTED_PARENT_ONLY_BATCHES = {"N3", "N6", "N7", "N8", "N11", "N13", "N14"}
EXPECTED_WORKBOOK_ONLY_VARIANTS = {
    "N12E",
    "N12F",
    "N10G",
    "N10P",
    "N6A",
    "N6B",
    "N6C",
    "N6D",
    "N6REMAKE",
    "N5B",
    "N5C",
    "N5D",
    "N5E",
    "N5H",
    "N5I",
    "N2AB2",
    "N2B",
}
PARENTLESS_VARIANT_KEYS = {"N16A", "N16B"}
OMITTED_PARENT_BATCHES = {"N16"}
USER_VARIANT_OVERRIDES = {
    "N10A": {
        "variant_descriptor": "Separate N10 child branch",
        "override_notes": "User-confirmed separate child of N10.",
        "metadata": {
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N10D": {
        "electrolyte": "1M LiTFSI + 10% FEC",
        "override_notes": "User-confirmed electrolyte override supplied outside the workbook.",
    },
    "N10E": {
        "variant_descriptor": "Separate N10 child branch",
        "override_notes": (
            "User response was ambiguous; importer currently treats N10e as a separate child of N10 "
            "to avoid incorrectly merging it with another branch."
        ),
        "metadata": {
            "lineage_annotation": "assumed_separate_child_branch",
        },
    },
    "N10N": {
        "variant_descriptor": "Separate N10 child branch",
        "override_notes": "User-confirmed separate child of N10.",
        "metadata": {
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N12B": {
        "variant_descriptor": "Hx-e (acid washed)",
        "post_processing": "acid_wash",
        "override_notes": "User-confirmed child of N12 using acid-washed Hx-e at higher loading.",
        "metadata": {
            "study_focus": "higher_loading",
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N9W": {
        "variant_descriptor": "Calendared porosity study",
        "override_notes": "User-confirmed child of N9 for calendared porosity evaluation.",
        "metadata": {
            "study_focus": "porosity",
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N9REMAKES": {
        "variant_descriptor": "Verification remake",
        "override_notes": "User-confirmed child of N9 created later to verify prior results.",
        "metadata": {
            "study_focus": "result_verification",
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N9X": {
        "variant_descriptor": "Calendared porosity study",
        "override_notes": "User-confirmed separate calendaring-study child of N9.",
        "metadata": {
            "study_focus": "porosity",
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N9Y": {
        "variant_descriptor": "Calendared porosity study",
        "override_notes": "User-confirmed separate calendaring-study child of N9.",
        "metadata": {
            "study_focus": "porosity",
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N1A": {
        "variant_descriptor": "Electrolyte evaluation",
        "override_notes": "User-confirmed child of N1 used for electrolyte evaluation.",
        "metadata": {
            "study_focus": "electrolyte_evaluation",
            "lineage_annotation": "user_confirmed_child_branch",
        },
    },
    "N16A": {
        "variant_descriptor": "Hx-e (concentrated HCl acid washed)",
        "post_processing": "concentrated_hcl_acid_wash",
    },
    "N16B": {
        "variant_descriptor": "Hx-e (conc. HCl acid wash and base washed)",
        "post_processing": "concentrated_hcl_then_base_wash",
    },
}

NEW_FORMAT_PARAMETER_FIELDS = {
    "disc mass (mg)": "disc_mass_mg",
    "electrode thickness (um)": "electrode_thickness_um",
    "pressed thickness (um)": "pressed_thickness_um",
    "active mass density, g/cc": "active_mass_density_g_cc",
    "slurry density (g/ml)": "slurry_density_g_ml",
    "electrode density (g/ml)": "electrode_density_g_ml",
    "porosity": "porosity",
    "thickness reduction": "thickness_reduction_pct",
}


def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_float(value: Any) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    for candidate in (text[:10], text):
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00")).date()
        except ValueError:
            continue
    return None


def _as_datetime_from_date(value: Optional[date]) -> Optional[datetime]:
    if value is None:
        return None
    return datetime.combine(value, datetime.min.time())


def _as_percent(value: Any) -> Optional[float]:
    number = _as_float(value)
    if number is None:
        return None
    return number * 100.0 if number <= 1.0 else number


def _normalize_key(value: Any) -> Optional[str]:
    text = _text(value)
    if not text:
        return None
    return re.sub(r"\s+", "", text).upper()


def _normalize_label(value: Any) -> Optional[str]:
    text = _text(value)
    if not text:
        return None
    normalized = text.lower()
    normalized = normalized.replace("µ", "u")
    normalized = normalized.replace("%", "pct")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip(" .")


def _cell_fill_hex(cell: Any) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    if getattr(fg, "type", None) == "rgb":
        rgb = _text(getattr(fg, "rgb", None))
        return rgb.upper() if rgb else None
    return None


def _is_review_fill(cell: Any) -> bool:
    return (_cell_fill_hex(cell) or "") in REVIEW_FILL_COLORS


def _extract_parent_code(name: str) -> Optional[str]:
    match = re.match(r"^(N\d+)", _text(name) or "", re.IGNORECASE)
    return match.group(1).upper() if match else None


def _canonical_variant_key(name: str) -> Optional[str]:
    raw = _text(name)
    if not raw:
        return None
    raw = re.sub(r"\s*-\s*repaired\b", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s+", "", raw)
    return raw.upper()


def _combine_notes(*values: Optional[str]) -> Optional[str]:
    parts: List[str] = []
    for value in values:
        text = _text(value)
        if text and text not in parts:
            parts.append(text)
    return "\n".join(parts) if parts else None


def _infer_calendared(*values: Optional[str]) -> Optional[bool]:
    combined = " ".join(filter(None, (_text(value) for value in values))).lower()
    if not combined:
        return None
    if "uncal" in combined:
        return False
    if re.search(r"\bcal\b", combined):
        return True
    if "calender" in combined or "calendar" in combined:
        return True
    return None


def _normalize_material_name(raw_name: str) -> Tuple[str, MaterialCategory, Dict[str, Any]]:
    text = _text(raw_name) or "Unknown"
    normalized = text.lower()
    normalized = normalized.replace("β", "beta")
    normalized = normalized.replace("(", " ").replace(")", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()

    metadata: Dict[str, Any] = {"source_name": text}
    if "nmc811" in normalized:
        if "basf" in normalized:
            metadata["manufacturer_hint"] = "BASF"
        elif "mse" in normalized:
            metadata["manufacturer_hint"] = "MSE"
        return "NMC811", MaterialCategory.CATHODE_ACTIVE, metadata
    if "pvdf" in normalized:
        metadata["manufacturer_hint"] = "Arkema"
        metadata["product_hint"] = "HSV1810"
        return "PVDF HSV1810", MaterialCategory.BINDER, metadata
    if normalized in {"al", "aluminum"}:
        return "Aluminum", MaterialCategory.CURRENT_COLLECTOR, metadata
    if "ccs" in normalized:
        return text, MaterialCategory.SEPARATOR, metadata
    if "super p" in normalized or "carbon black" in normalized:
        return "Super P", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "mwcnt" in normalized:
        return "MWCNTs", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "cx-e" in normalized or "cx echo" in normalized or "cx e" in normalized or "hx-t" in normalized:
        metadata["manufacturer_hint"] = "Hexegen"
        return "Hx-T", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if (
        "cx b" in normalized
        or "cx beta" in normalized
        or "hx-e" in normalized
        or "hx e" in normalized
    ):
        metadata["manufacturer_hint"] = "Hexegen"
        treatment = _extract_treatment_hint(text)
        if treatment:
            metadata["post_processing"] = treatment
        return "Hx-e", MaterialCategory.CONDUCTIVE_ADDITIVE, metadata
    if "lipaa" in normalized:
        return "LiPAA", MaterialCategory.BINDER, metadata
    if normalized == "sbr":
        return "SBR", MaterialCategory.BINDER, metadata
    if normalized == "lfp":
        return "LFP", MaterialCategory.OTHER, metadata
    return text, MaterialCategory.OTHER, metadata


def _extract_treatment_hint(text: str) -> Optional[str]:
    normalized = (_text(text) or "").lower()
    if not normalized:
        return None
    if "conc" in normalized and "base" in normalized:
        return "concentrated_hcl_then_base_wash"
    if "conc" in normalized and ("acid" in normalized or "hcl" in normalized):
        return "concentrated_hcl_acid_wash"
    if "base" in normalized:
        return "acid_then_base_wash"
    if "acid" in normalized or "hcl" in normalized:
        return "acid_wash"
    return None


def _extract_branch_descriptors(note_text: Optional[str]) -> Dict[str, str]:
    note = _text(note_text) or ""
    descriptors: Dict[str, str] = {}
    pattern = re.compile(
        r"(?P<code>N\d+[A-Za-z0-9]+)\s*:\s*(?P<desc>.*?)(?=(?:\band\b\s+N\d+[A-Za-z0-9]+\s*:)|\)|$)",
        re.IGNORECASE,
    )
    for match in pattern.finditer(note):
        code = _canonical_variant_key(match.group("code"))
        desc = _text(match.group("desc"))
        if code and desc:
            descriptors[code] = desc.rstrip(" -")
    return descriptors


def _descriptor_to_component_metadata(descriptor: Optional[str]) -> Dict[str, Any]:
    text = _text(descriptor)
    if not text:
        return {}
    metadata = {"variant_descriptor": text}
    treatment = _extract_treatment_hint(text)
    if treatment:
        metadata["post_processing"] = treatment
    return metadata


def _normalize_electrolyte_candidates(value: Optional[str]) -> List[str]:
    text = _text(value)
    if not text:
        return []
    candidates = []
    for part in text.split(";"):
        candidate = re.sub(r"[^a-z0-9]+", "", part.lower())
        if candidate:
            candidates.append(candidate)
    return candidates


def _electrolyte_token_set(value: Optional[str]) -> set[str]:
    text = (_text(value) or "").lower()
    if not text:
        return set()
    normalized = text
    normalized = normalized.replace("%", "pct")
    normalized = normalized.replace(":", " ")
    normalized = normalized.replace("+", " ")
    normalized = normalized.replace(",", " ")
    normalized = normalized.replace("(", " ")
    normalized = normalized.replace(")", " ")
    tokens = set(re.findall(r"[a-z0-9\.]+", normalized))
    return {token for token in tokens if token not in {"", "ec", "emc", "dmc", "tegdme"}}


def _electrolytes_match(workbook_value: Optional[str], legacy_value: Optional[str]) -> bool:
    workbook_candidates = _normalize_electrolyte_candidates(workbook_value)
    legacy_candidates = _normalize_electrolyte_candidates(legacy_value)
    if not workbook_candidates or not legacy_candidates:
        return workbook_value == legacy_value
    for workbook_candidate in workbook_candidates:
        for legacy_candidate in legacy_candidates:
            if (
                workbook_candidate == legacy_candidate
                or workbook_candidate in legacy_candidate
                or legacy_candidate in workbook_candidate
            ):
                return True
    workbook_tokens = _electrolyte_token_set(workbook_value)
    legacy_tokens = _electrolyte_token_set(legacy_value)
    if workbook_tokens and legacy_tokens:
        if workbook_tokens <= legacy_tokens or legacy_tokens <= workbook_tokens:
            return True
    return False


def _flatten_issue_rows(issue_type: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    flattened: List[Dict[str, Any]] = []
    for item in items:
        flattened.append(
            {
                "issue_type": issue_type,
                "status": item.get("status"),
                "severity": item.get("severity"),
                "scope": item.get("scope"),
                "record_name": item.get("record_name"),
                "field": item.get("field"),
                "message": item.get("message"),
                "workbook_value": json.dumps(item.get("workbook_value"), default=str),
                "legacy_value": json.dumps(item.get("legacy_value"), default=str),
                "workbook_ref": item.get("workbook_ref"),
                "legacy_ref": item.get("legacy_ref"),
            }
        )
    return flattened


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")


def _write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_markdown_summary(path: Path, plan: Dict[str, Any]) -> None:
    counts = plan["counts"]
    lines = [
        "# NMC Cathode Ontology Reconciliation",
        "",
        f"- Source workbook: `{plan['source_workbook']}`",
        f"- Legacy database: `{plan['legacy_db_path']}`",
        f"- Comparison workbook: `{plan.get('comparison_workbook_path') or 'None'}`",
        "",
        "## Counts",
        "",
        f"- Parent batches: {counts['parent_batches']}",
        f"- Variant batches: {counts['variant_batches']}",
        f"- Materials: {counts['materials']}",
        f"- Review items: {counts['review_items']}",
        f"- Mismatch items: {counts['mismatch_items']}",
        "",
        "## What Needs Reconciliation",
        "",
        "- Confirm highlighted parent-sheet inputs that were intentionally skipped.",
        "- Resolve workbook-only and legacy-only experiment aliases before persistence.",
        "- Review real data disagreements where workbook and DB both exist.",
        "",
        "## Files",
        "",
        "- `ontology_import_preview.json`: full preview payload",
        "- `ontology_import_issues.csv`: flattened review + mismatch issues",
        "- `ontology_import_parent_batches.json`: parent batch preview",
        "- `ontology_import_variant_batches.json`: variant batch preview",
        "- `ontology_import_cell_builds.json`: proposed cell-build lineage preview",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_review_export(plan: Dict[str, Any], export_dir: Path) -> Dict[str, str]:
    export_dir.mkdir(parents=True, exist_ok=True)

    preview_path = export_dir / "ontology_import_preview.json"
    issues_path = export_dir / "ontology_import_issues.csv"
    parent_batches_path = export_dir / "ontology_import_parent_batches.json"
    variant_batches_path = export_dir / "ontology_import_variant_batches.json"
    cell_builds_path = export_dir / "ontology_import_cell_builds.json"
    readme_path = export_dir / "README.md"

    _write_json(preview_path, plan)
    _write_json(parent_batches_path, plan["parent_batches"])
    _write_json(variant_batches_path, plan["variant_batches"])
    _write_json(cell_builds_path, plan["cell_builds"])
    _write_csv(
        issues_path,
        [
            *_flatten_issue_rows("review_required", plan["review_items"]),
            *_flatten_issue_rows("mismatch", plan["mismatch_items"]),
        ],
    )
    _write_markdown_summary(readme_path, plan)

    return {
        "export_dir": str(export_dir),
        "preview_json": str(preview_path),
        "issues_csv": str(issues_path),
        "parent_batches_json": str(parent_batches_path),
        "variant_batches_json": str(variant_batches_path),
        "cell_builds_json": str(cell_builds_path),
        "readme_md": str(readme_path),
    }


def _serialize_component(component: Dict[str, Any]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "Component": component["Component"],
        "Dry Mass Fraction (%)": round(component["Dry Mass Fraction (%)"], 4),
    }
    metadata = component.get("metadata") or {}
    if metadata:
        payload["metadata"] = metadata
    return payload


def _component_signature(components: Iterable[Dict[str, Any]]) -> List[Tuple[str, float, Optional[str]]]:
    signature = []
    for component in components:
        metadata = component.get("metadata") or {}
        signature.append(
            (
                component["Component"],
                round(float(component["Dry Mass Fraction (%)"]), 2),
                metadata.get("post_processing"),
            )
        )
    return sorted(signature)


def _numbers_close(left: Optional[float], right: Optional[float], tolerance: float = 0.11) -> bool:
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def _create_issue(
    *,
    status: str,
    scope: str,
    record_name: str,
    field: str,
    message: str,
    workbook_value: Any = None,
    legacy_value: Any = None,
    workbook_ref: Optional[str] = None,
    legacy_ref: Optional[str] = None,
    severity: str = "warning",
) -> Dict[str, Any]:
    return {
        "status": status,
        "severity": severity,
        "scope": scope,
        "record_name": record_name,
        "field": field,
        "message": message,
        "workbook_value": workbook_value,
        "legacy_value": legacy_value,
        "workbook_ref": workbook_ref,
        "legacy_ref": legacy_ref,
    }


def _nmc_source_for_parent(parent_code: str) -> str:
    number = int(parent_code[1:])
    return "MSE" if number <= 7 or parent_code == "N11" else "BASF"


def _extract_sheet_measurements(value_ws: Any, style_ws: Any, parent_code: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    safe_metrics: Dict[str, Any] = {}
    review_items: List[Dict[str, Any]] = []

    for row in range(29, 37):
        field_key = NEW_FORMAT_PARAMETER_FIELDS.get(_normalize_label(value_ws[f"B{row}"].value) or "")
        if not field_key:
            continue
        value_cell = value_ws[f"C{row}"]
        style_cell = style_ws[f"C{row}"]
        value = _as_float(value_cell.value)
        if value is None:
            continue
        if _is_review_fill(style_cell):
            review_items.append(
                _create_issue(
                    status="review_required",
                    scope="parent_sheet",
                    record_name=parent_code,
                    field=field_key,
                    message="Highlighted workbook value was skipped and needs manual confirmation.",
                    workbook_value=value,
                    workbook_ref=f"{value_ws.title}!{value_cell.coordinate}",
                )
            )
            continue
        safe_metrics[field_key] = value

    if "electrode_thickness_um" not in safe_metrics and _text(value_ws["D23"].value):
        thickness = _as_float(value_ws["D24"].value)
        if thickness is not None:
            if _is_review_fill(style_ws["D24"]):
                review_items.append(
                    _create_issue(
                        status="review_required",
                        scope="parent_sheet",
                        record_name=parent_code,
                        field="electrode_thickness_um",
                        message="Highlighted workbook value was skipped and needs manual confirmation.",
                        workbook_value=thickness,
                        workbook_ref=f"{value_ws.title}!D24",
                    )
                )
            else:
                safe_metrics["electrode_thickness_um"] = thickness

    if "pressed_thickness_um" not in safe_metrics and _text(value_ws["E23"].value):
        pressed = _as_float(value_ws["E24"].value)
        if pressed is not None:
            if _is_review_fill(style_ws["E24"]):
                review_items.append(
                    _create_issue(
                        status="review_required",
                        scope="parent_sheet",
                        record_name=parent_code,
                        field="pressed_thickness_um",
                        message="Highlighted workbook value was skipped and needs manual confirmation.",
                        workbook_value=pressed,
                        workbook_ref=f"{value_ws.title}!E24",
                    )
                )
            else:
                safe_metrics["pressed_thickness_um"] = pressed

    if "disc_mass_mg" not in safe_metrics and _normalize_label(value_ws["E17"].value) == "disc mass (mg)":
        disc_mass = _as_float(value_ws["E18"].value)
        if disc_mass is not None:
            if _is_review_fill(style_ws["E18"]):
                review_items.append(
                    _create_issue(
                        status="review_required",
                        scope="parent_sheet",
                        record_name=parent_code,
                        field="disc_mass_mg",
                        message="Highlighted workbook value was skipped and needs manual confirmation.",
                        workbook_value=disc_mass,
                        workbook_ref=f"{value_ws.title}!E18",
                    )
                )
            else:
                safe_metrics["disc_mass_mg"] = disc_mass

    return safe_metrics, review_items


def _parse_parent_sheets(workbook_path: Path) -> Dict[str, Dict[str, Any]]:
    workbook_values = load_workbook(workbook_path, data_only=True)
    workbook_styles = load_workbook(workbook_path, data_only=False)

    parents: Dict[str, Dict[str, Any]] = {}
    for sheet_name in workbook_values.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue
        ws_values = workbook_values[sheet_name]
        ws_styles = workbook_styles[sheet_name]
        parent_code = (_text(ws_values["B1"].value) or sheet_name).upper()
        notes = _text(ws_values["D1"].value)
        result_notes = _text(ws_values["F1"].value)
        branch_descriptors = _extract_branch_descriptors(notes)
        has_branch_specific_descriptors = len(branch_descriptors) > 1

        formulation: List[Dict[str, Any]] = []
        for row in range(4, 12):
            raw_name = _text(ws_values.cell(row, 1).value)
            fraction_pct = _as_percent(ws_values.cell(row, 2).value)
            if not raw_name or fraction_pct is None or fraction_pct <= 0:
                continue
            canonical_name, category, metadata = _normalize_material_name(raw_name)
            if has_branch_specific_descriptors and canonical_name == "Hx-e":
                metadata = {k: v for k, v in metadata.items() if k != "post_processing"}
            formulation.append(
                {
                    "Component": canonical_name,
                    "Dry Mass Fraction (%)": fraction_pct,
                    "category": category,
                    "metadata": metadata,
                }
            )

        parent_treatment = _extract_treatment_hint(notes or "")
        if parent_treatment and not has_branch_specific_descriptors:
            for component in formulation:
                if component["Component"] == "Hx-e":
                    component.setdefault("metadata", {})["post_processing"] = parent_treatment

        safe_metrics, review_items = _extract_sheet_measurements(ws_values, ws_styles, parent_code)
        parents[parent_code] = {
            "batch_name": parent_code,
            "build_date": _as_date(ws_values["B2"].value),
            "notes": notes,
            "result_notes": result_notes,
            "branch_descriptors": branch_descriptors,
            "active_material_source": _nmc_source_for_parent(parent_code),
            "formulation": formulation,
            "safe_metrics": safe_metrics,
            "review_items": review_items,
            "source_sheet": sheet_name,
        }
    return parents


def _parse_summary_rows(workbook_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    workbook = load_workbook(workbook_path, data_only=True)
    ws = workbook[SUMMARY_SHEET]
    rows_by_parent: Dict[str, List[Dict[str, Any]]] = {}
    for row in range(2, ws.max_row + 1):
        raw_name = _text(ws.cell(row, 2).value)
        parent_code = _extract_parent_code(raw_name or "")
        if not raw_name or not parent_code:
            continue
        match_key = _canonical_variant_key(raw_name)
        notes = _text(ws.cell(row, 19).value)
        record = {
            "row_index": row,
            "raw_name": raw_name,
            "match_key": match_key,
            "parent_code": parent_code,
            "is_exact_parent": match_key == parent_code,
            "date": _as_date(ws.cell(row, 1).value),
            "summary_active_material_pct": _as_percent(ws.cell(row, 3).value),
            "summary_pvdf_pct": _as_percent(ws.cell(row, 4).value),
            "summary_cx_pct": _as_percent(ws.cell(row, 5).value),
            "summary_super_p_pct": _as_percent(ws.cell(row, 6).value),
            "summary_cnt_pct": _as_percent(ws.cell(row, 7).value),
            "solids_content_pct": _as_percent(ws.cell(row, 8).value),
            "mixing_strategy": _text(ws.cell(row, 9).value),
            "current_collector": _text(ws.cell(row, 10).value),
            "electrolyte": _text(ws.cell(row, 11).value),
            "cycler_channels": _text(ws.cell(row, 16).value),
            "cycling_strategy": _text(ws.cell(row, 17).value),
            "results": _text(ws.cell(row, 18).value),
            "notes": notes,
            "calendared": _infer_calendared(notes),
            "workbook_ref": f"{SUMMARY_SHEET}!{row}",
        }
        rows_by_parent.setdefault(parent_code, []).append(record)
    return rows_by_parent


def _parse_comparison_workbook_rows(comparison_workbook_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    workbook = load_workbook(comparison_workbook_path, data_only=True)
    rows_by_parent: Dict[str, List[Dict[str, Any]]] = {}

    if "Compiled" in workbook.sheetnames:
        ws = workbook["Compiled"]
        for row in range(2, ws.max_row + 1):
            raw_name = _text(ws.cell(row, 1).value)
            parent_code = _extract_parent_code(raw_name or "")
            if not raw_name or not parent_code:
                continue
            notes = _text(ws.cell(row, 10).value)
            record = {
                "row_index": row,
                "raw_name": raw_name,
                "match_key": _canonical_variant_key(raw_name),
                "parent_code": parent_code,
                "is_exact_parent": _canonical_variant_key(raw_name) == parent_code,
                "date": None,
                "summary_active_material_pct": None,
                "summary_pvdf_pct": None,
                "summary_cx_pct": None,
                "summary_super_p_pct": None,
                "summary_cnt_pct": None,
                "solids_content_pct": None,
                "mixing_strategy": None,
                "current_collector": None,
                "electrolyte": _text(ws.cell(row, 8).value),
                "cycler_channels": None,
                "cycling_strategy": None,
                "results": _text(ws.cell(row, 9).value),
                "notes": notes,
                "calendared": _infer_calendared(notes),
                "workbook_ref": f"Compiled!{row}",
                "comparison_loading_mg": _as_float(ws.cell(row, 2).value),
                "comparison_active_mass_mg": _as_float(ws.cell(row, 3).value),
                "comparison_specific_discharge_mAh_g": _as_float(ws.cell(row, 6).value),
                "comparison_fce": _as_float(ws.cell(row, 7).value),
                "comparison_carbon_material": None,
            }
            rows_by_parent.setdefault(parent_code, []).append(record)

    if "Summary" in workbook.sheetnames:
        ws = workbook["Summary"]
        for row in range(2, ws.max_row + 1):
            raw_name = _text(ws.cell(row, 1).value)
            parent_code = _extract_parent_code(raw_name or "")
            if not raw_name or not parent_code:
                continue
            notes = _text(ws.cell(row, 6).value)
            record = {
                "row_index": row,
                "raw_name": raw_name,
                "match_key": _canonical_variant_key(raw_name),
                "parent_code": parent_code,
                "is_exact_parent": _canonical_variant_key(raw_name) == parent_code,
                "date": None,
                "summary_active_material_pct": None,
                "summary_pvdf_pct": None,
                "summary_cx_pct": None,
                "summary_super_p_pct": None,
                "summary_cnt_pct": None,
                "solids_content_pct": None,
                "mixing_strategy": None,
                "current_collector": None,
                "electrolyte": _text(ws.cell(row, 5).value),
                "cycler_channels": None,
                "cycling_strategy": None,
                "results": _text(ws.cell(row, 7).value),
                "notes": notes,
                "calendared": _infer_calendared(notes),
                "workbook_ref": f"Summary!{row}",
                "comparison_loading_mg": None,
                "comparison_active_mass_mg": None,
                "comparison_specific_discharge_mAh_g": _as_float(ws.cell(row, 2).value),
                "comparison_fce": _as_float(ws.cell(row, 3).value),
                "comparison_carbon_material": _text(ws.cell(row, 4).value),
            }
            rows_by_parent.setdefault(parent_code, []).append(record)

    return rows_by_parent


def _merge_authoritative_rows(
    primary_rows_by_parent: Dict[str, List[Dict[str, Any]]],
    comparison_rows_by_parent: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, List[Dict[str, Any]]]:
    merged_rows: Dict[str, List[Dict[str, Any]]] = {}
    for parent_code in set(primary_rows_by_parent) | set(comparison_rows_by_parent):
        rows_by_key: Dict[str, Dict[str, Any]] = {}
        for row in primary_rows_by_parent.get(parent_code, []):
            rows_by_key[row["match_key"]] = dict(row)

        for comparison_row in comparison_rows_by_parent.get(parent_code, []):
            row = rows_by_key.get(comparison_row["match_key"])
            if row is None:
                row = dict(comparison_row)
                row["source_origin"] = "comparison_workbook"
                rows_by_key[comparison_row["match_key"]] = row
                continue

            if comparison_row.get("electrolyte"):
                row["electrolyte"] = comparison_row["electrolyte"]
            if comparison_row.get("notes"):
                row["notes"] = _combine_notes(row.get("notes"), comparison_row.get("notes"))
            if comparison_row.get("results"):
                row["results"] = _combine_notes(row.get("results"), comparison_row.get("results"))
            if comparison_row.get("calendared") is not None:
                row["calendared"] = comparison_row["calendared"]
            if comparison_row.get("comparison_carbon_material"):
                row["comparison_carbon_material"] = comparison_row["comparison_carbon_material"]
            if comparison_row.get("comparison_specific_discharge_mAh_g") is not None:
                row["comparison_specific_discharge_mAh_g"] = comparison_row["comparison_specific_discharge_mAh_g"]
            if comparison_row.get("comparison_fce") is not None:
                row["comparison_fce"] = comparison_row["comparison_fce"]
            row["comparison_ref"] = comparison_row.get("workbook_ref")

        merged_rows[parent_code] = list(rows_by_key.values())

    return merged_rows


def _apply_row_overrides(rows_by_parent: Dict[str, List[Dict[str, Any]]]) -> Dict[str, List[Dict[str, Any]]]:
    for override_key, override in USER_VARIANT_OVERRIDES.items():
        parent_code = _extract_parent_code(override_key)
        if not parent_code:
            continue
        for row in rows_by_parent.get(parent_code, []):
            if row.get("match_key") != override_key:
                continue
            if override.get("electrolyte"):
                row["electrolyte"] = override["electrolyte"]
            if override.get("override_notes"):
                row["notes"] = _combine_notes(row.get("notes"), override["override_notes"])
            row["user_override_applied"] = True
    return rows_by_parent


def _normalize_legacy_formulation(raw_json: Optional[str]) -> List[Dict[str, Any]]:
    if not raw_json:
        return []
    try:
        data = json.loads(raw_json)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []

    parsed_components: List[Tuple[str, float]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        raw_name = _text(item.get("Component") or item.get("component"))
        raw_fraction = _as_float(
            item.get("Dry Mass Fraction (%)")
            or item.get("dry_mass_fraction_pct")
            or item.get("fraction")
        )
        if not raw_name or raw_fraction is None or raw_fraction <= 0:
            continue
        parsed_components.append((raw_name, raw_fraction))

    if not parsed_components:
        return []

    raw_total = sum(value for _, value in parsed_components)
    scale = 100.0 if raw_total <= 2.0 else 1.0

    normalized: List[Dict[str, Any]] = []
    for raw_name, raw_fraction in parsed_components:
        fraction_pct = raw_fraction * scale
        canonical_name, _, metadata = _normalize_material_name(raw_name)
        normalized.append(
            {
                "Component": canonical_name,
                "Dry Mass Fraction (%)": fraction_pct,
                "metadata": metadata,
            }
        )
    return normalized


def _load_legacy_records(legacy_db_path: Path) -> Dict[str, Any]:
    conn = sqlite3.connect(legacy_db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT
                ce.id,
                ce.project_id,
                p.name AS project_name,
                p.project_type,
                ce.cell_name,
                ce.loading,
                ce.active_material,
                ce.electrolyte,
                ce.formulation_json,
                ce.solids_content,
                ce.pressed_thickness,
                ce.experiment_notes,
                ce.substrate,
                ce.separator,
                ce.data_json
            FROM cell_experiments ce
            JOIN projects p ON p.id = ce.project_id
            WHERE p.name IN ('NMC Half Cells', 'NMC- Si Full Cell')
            ORDER BY p.name, ce.cell_name
            """
        ).fetchall()
    finally:
        conn.close()

    half_cells_by_parent: Dict[str, List[Dict[str, Any]]] = {}
    full_cells_by_name: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        data_json: Dict[str, Any] = {}
        try:
            if row["data_json"]:
                loaded = json.loads(row["data_json"])
                if isinstance(loaded, dict):
                    data_json = loaded
        except json.JSONDecodeError:
            data_json = {}

        legacy_record = {
            "id": row["id"],
            "project_id": row["project_id"],
            "project_name": row["project_name"],
            "project_type": row["project_type"],
            "cell_name": row["cell_name"],
            "match_key": _canonical_variant_key(row["cell_name"]),
            "parent_code": _extract_parent_code(row["cell_name"] or ""),
            "loading": _as_float(row["loading"]),
            "active_material_pct": _as_percent(row["active_material"]),
            "electrolyte": _text(row["electrolyte"]),
            "formulation": _normalize_legacy_formulation(row["formulation_json"]),
            "solids_content_pct": _as_percent(row["solids_content"]),
            "pressed_thickness_um": _as_float(row["pressed_thickness"]),
            "notes": _text(row["experiment_notes"]),
            "substrate": _text(row["substrate"]),
            "separator": _text(row["separator"]),
            "data_json": data_json,
            "calendared": _infer_calendared(_text(row["experiment_notes"]), _text(data_json.get("tracking"))),
            "is_repaired": "repaired" in (_text(row["cell_name"]) or "").lower(),
        }

        if legacy_record["project_name"] == "NMC Half Cells" and legacy_record["parent_code"]:
            half_cells_by_parent.setdefault(legacy_record["parent_code"], []).append(legacy_record)
        else:
            full_cells_by_name[legacy_record["match_key"]] = legacy_record

    return {
        "half_cells_by_parent": half_cells_by_parent,
        "full_cells_by_name": full_cells_by_name,
    }


def _collect_legacy_loading_values(records: List[Dict[str, Any]]) -> List[float]:
    values: List[float] = []
    for record in records:
        record_values: List[float] = []
        cells = record.get("data_json", {}).get("cells")
        if isinstance(cells, list):
            for cell in cells:
                if not isinstance(cell, dict):
                    continue
                value = _as_float(cell.get("loading"))
                if value is not None:
                    record_values.append(value)
        top_level = _as_float(record.get("loading"))
        if top_level is not None and not record_values:
            record_values.append(top_level)
        values.extend(record_values)
    return values


def _derive_stable_legacy_loading(
    records: List[Dict[str, Any]],
    *,
    tolerance: float = 0.2,
) -> Optional[Dict[str, Any]]:
    loading_values = _collect_legacy_loading_values(records)
    if not loading_values:
        return None

    min_value = min(loading_values)
    max_value = max(loading_values)
    if max_value - min_value > tolerance:
        return None

    representative_value = round(sum(loading_values) / len(loading_values), 4)
    return {
        "value": representative_value,
        "sample_count": len(loading_values),
        "record_ids": [record["id"] for record in records],
        "source": "legacy_cell_loadings",
    }


def _choose_primary_legacy_record(records: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not records:
        return None
    repaired = [record for record in records if record["is_repaired"]]
    candidates = repaired or records
    return sorted(candidates, key=lambda record: (record["id"], record["cell_name"]))[-1]


def _build_variant_formulation(
    parent_formulation: List[Dict[str, Any]],
    *,
    descriptor: Optional[str] = None,
    legacy_override: Optional[List[Dict[str, Any]]] = None,
    user_override: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    if legacy_override:
        formulation = [_serialize_component(component) for component in legacy_override]
    else:
        formulation = copy.deepcopy(parent_formulation)
        descriptor_metadata = _descriptor_to_component_metadata(descriptor)
        if descriptor_metadata:
            for component in formulation:
                if component["Component"] == "Hx-e":
                    component.setdefault("metadata", {}).update(descriptor_metadata)
        formulation = [_serialize_component(component) for component in formulation]

    if user_override:
        override_metadata = {}
        if user_override.get("variant_descriptor"):
            override_metadata["variant_descriptor"] = user_override["variant_descriptor"]
        if user_override.get("post_processing"):
            override_metadata["post_processing"] = user_override["post_processing"]
        for component in formulation:
            if component["Component"] == "Hx-e":
                component.setdefault("metadata", {}).update(override_metadata)

    return formulation


def _apply_batch_material_hints(
    formulation: List[Dict[str, Any]],
    *,
    active_material_source: Optional[str],
) -> List[Dict[str, Any]]:
    enriched = copy.deepcopy(formulation)
    for component in enriched:
        if not isinstance(component, dict):
            continue
        component_name = _text(component.get("Component"))
        metadata = component.setdefault("metadata", {})
        if component_name == "NMC811" and active_material_source:
            metadata["manufacturer_hint"] = active_material_source
        if component_name in {"Hx-e", "Hx-T"} and not metadata.get("manufacturer_hint"):
            metadata["manufacturer_hint"] = "Hexegen"
    return enriched


def _merge_metadata_dicts(*values: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    merged: Dict[str, Any] = {}
    for value in values:
        if not value:
            continue
        merged.update(value)
    return merged or None


def _add_material_spec(materials: Dict[str, Dict[str, Any]], name: str, category: MaterialCategory, metadata: Optional[Dict[str, Any]] = None) -> None:
    existing = materials.get(name)
    if existing:
        merged = {**(existing.get("metadata_json") or {}), **(metadata or {})}
        existing["metadata_json"] = merged or None
        return
    materials[name] = {
        "name": name,
        "category": category,
        "manufacturer": (metadata or {}).get("manufacturer_hint"),
        "description": None,
        "metadata_json": metadata or None,
    }


def _compare_summary_to_legacy(
    *,
    record_name: str,
    scope: str,
    summary_row: Optional[Dict[str, Any]],
    legacy_record: Optional[Dict[str, Any]],
    expected_formulation: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    if not summary_row or not legacy_record:
        return issues

    if not _electrolytes_match(summary_row.get("electrolyte"), legacy_record.get("electrolyte")):
        issues.append(
            _create_issue(
                status="mismatch",
                scope=scope,
                record_name=record_name,
                field="electrolyte",
                message="Workbook electrolyte does not match the existing database entry.",
                workbook_value=summary_row.get("electrolyte"),
                legacy_value=legacy_record.get("electrolyte"),
                workbook_ref=summary_row.get("workbook_ref"),
                legacy_ref=f"cell_experiments:{legacy_record['id']}",
            )
        )

    workbook_solids = summary_row.get("solids_content_pct")
    legacy_solids = legacy_record.get("solids_content_pct")
    if (
        workbook_solids is not None
        and legacy_solids is not None
        and not _numbers_close(workbook_solids, legacy_solids)
    ):
        issues.append(
            _create_issue(
                status="mismatch",
                scope=scope,
                record_name=record_name,
                field="solids_content_pct",
                message="Workbook solids content does not match the existing database entry.",
                workbook_value=workbook_solids,
                legacy_value=legacy_solids,
                workbook_ref=summary_row.get("workbook_ref"),
                legacy_ref=f"cell_experiments:{legacy_record['id']}",
            )
        )

    workbook_active = summary_row.get("summary_active_material_pct")
    legacy_active = legacy_record.get("active_material_pct")
    if (
        workbook_active is not None
        and legacy_active is not None
        and not _numbers_close(workbook_active, legacy_active)
    ):
        issues.append(
            _create_issue(
                status="mismatch",
                scope=scope,
                record_name=record_name,
                field="active_material_pct",
                message="Workbook active-material percentage does not match the existing database entry.",
                workbook_value=workbook_active,
                legacy_value=legacy_active,
                workbook_ref=summary_row.get("workbook_ref"),
                legacy_ref=f"cell_experiments:{legacy_record['id']}",
            )
        )

    legacy_formulation = legacy_record.get("formulation") or []
    if legacy_formulation and _component_signature(expected_formulation) != _component_signature(legacy_formulation):
        issues.append(
            _create_issue(
                status="mismatch",
                scope=scope,
                record_name=record_name,
                field="formulation",
                message="Workbook-derived formulation does not match the existing database entry.",
                workbook_value=expected_formulation,
                legacy_value=legacy_formulation,
                workbook_ref=summary_row.get("workbook_ref"),
                legacy_ref=f"cell_experiments:{legacy_record['id']}",
            )
        )
    return issues


def _create_cell_build_specs(
    *,
    source_batch_name: str,
    legacy_record: Dict[str, Any],
    relationship_type: str,
    relationship_notes: Optional[str] = None,
) -> List[Dict[str, Any]]:
    cells = legacy_record.get("data_json", {}).get("cells")
    if not isinstance(cells, list) or not cells:
        return []

    tracking = legacy_record.get("data_json", {}).get("tracking")
    tracking_date = None
    if isinstance(tracking, dict):
        tracking_date = _as_date(tracking.get("tracking_date"))
    build_date = _as_date(legacy_record.get("data_json", {}).get("experiment_date")) or tracking_date

    build_specs: List[Dict[str, Any]] = []
    for index, cell in enumerate(cells, start=1):
        if not isinstance(cell, dict):
            continue
        build_name = _text(cell.get("cell_name")) or _text(cell.get("test_number")) or f"{legacy_record['cell_name']}::{index}"
        metadata = {
            "legacy_project_name": legacy_record["project_name"],
            "legacy_cell_name": legacy_record["cell_name"],
            "legacy_test_number": _text(cell.get("test_number")),
            "electrolyte": _text(cell.get("electrolyte")) or legacy_record.get("electrolyte"),
            "separator": _text(cell.get("separator")) or legacy_record.get("separator"),
            "substrate": _text(cell.get("substrate")) or legacy_record.get("substrate"),
            "loading": _as_float(cell.get("loading")) or legacy_record.get("loading"),
            "active_material_pct": _as_percent(cell.get("active_material")) or legacy_record.get("active_material_pct"),
            "source_batch_name": source_batch_name,
            "tracking_date": tracking_date.isoformat() if tracking_date else None,
        }
        build_specs.append(
            {
                "build_name": build_name,
                "chemistry": "NMC811 half-cell" if legacy_record["project_name"] == "NMC Half Cells" else "NMC-Si full cell",
                "form_factor": None,
                "build_date": build_date,
                "status": CellBuildStatus.TESTING,
                "notes": _combine_notes(legacy_record.get("notes"), relationship_notes),
                "legacy_project_id": legacy_record["project_id"],
                "legacy_experiment_id": legacy_record["id"],
                "legacy_cell_id": None,
                "legacy_test_number": _text(cell.get("test_number")),
                "metadata_json": metadata,
                "source_batch_name": source_batch_name,
                "relationship_type": relationship_type,
                "relationship_notes": relationship_notes,
            }
        )
    return build_specs


def _prepare_import_plan(
    workbook_path: Path,
    legacy_db_path: Path,
    comparison_workbook_path: Optional[Path] = None,
) -> Dict[str, Any]:
    parents = _parse_parent_sheets(workbook_path)
    summary_rows_by_parent = _parse_summary_rows(workbook_path)
    if comparison_workbook_path:
        comparison_rows_by_parent = _parse_comparison_workbook_rows(comparison_workbook_path)
        summary_rows_by_parent = _merge_authoritative_rows(summary_rows_by_parent, comparison_rows_by_parent)
    summary_rows_by_parent = _apply_row_overrides(summary_rows_by_parent)
    legacy = _load_legacy_records(legacy_db_path)

    materials: Dict[str, Dict[str, Any]] = {}
    protocols: Dict[Tuple[str, str], Dict[str, Any]] = {}
    process_runs: List[Dict[str, Any]] = []
    parent_batches: List[Dict[str, Any]] = []
    variant_batches: List[Dict[str, Any]] = []
    experiment_links: List[Dict[str, Any]] = []
    cell_builds: List[Dict[str, Any]] = []
    review_items: List[Dict[str, Any]] = []
    mismatch_items: List[Dict[str, Any]] = []
    notes: List[str] = [
        "Hx-e acid-wash and acid+base-wash variants are stored as component metadata, not separate canonical materials.",
        "NMC811 source is preserved in batch metadata as MSE or BASF until lot numbers are available.",
    ]

    for parent_code, parent in parents.items():
        summary_rows = summary_rows_by_parent.get(parent_code, [])
        legacy_rows = legacy["half_cells_by_parent"].get(parent_code, [])
        exact_summary = next((row for row in summary_rows if row["is_exact_parent"]), None)
        exact_legacy_records = [row for row in legacy_rows if row["match_key"] == parent_code]
        exact_legacy = _choose_primary_legacy_record(exact_legacy_records)
        stable_parent_loading = None
        if "disc_mass_mg" not in parent["safe_metrics"]:
            stable_parent_loading = _derive_stable_legacy_loading(legacy_rows)
            if stable_parent_loading:
                parent["safe_metrics"]["disc_mass_mg"] = stable_parent_loading["value"]
                parent["review_items"] = [
                    item
                    for item in parent["review_items"]
                    if item.get("field") != "disc_mass_mg"
                ]
                notes.append(
                    f"{parent_code} disc mass was backfilled from stable legacy loading data "
                    f"({stable_parent_loading['value']} mg across {stable_parent_loading['sample_count']} cell(s))."
                )

        review_items.extend(parent["review_items"])
        for component in parent["formulation"]:
            _add_material_spec(
                materials,
                component["Component"],
                component["category"],
                component.get("metadata"),
            )

        current_collector = None
        if exact_summary and exact_summary.get("current_collector"):
            current_collector_name, category, metadata = _normalize_material_name(exact_summary["current_collector"])
            _add_material_spec(materials, current_collector_name, category, metadata)
            current_collector = current_collector_name

        separator = exact_legacy.get("separator") if exact_legacy else None
        if separator:
            separator_name, category, metadata = _normalize_material_name(separator)
            _add_material_spec(materials, separator_name, category, metadata)
            separator = separator_name

        mixing_strategy = exact_summary.get("mixing_strategy") if exact_summary else None
        if not mixing_strategy:
            for row in summary_rows:
                if row.get("mixing_strategy"):
                    mixing_strategy = row["mixing_strategy"]
                    break
        protocol_key = None
        if mixing_strategy:
            protocol_key = (mixing_strategy, IMPORT_VERSION)
            protocols[protocol_key] = {
                "name": mixing_strategy,
                "version": IMPORT_VERSION,
                "protocol_type": ProtocolType.OTHER,
                "description": f"Imported from {workbook_path.name} summary sheet.",
                "step_definition_json": None,
                "metadata_json": {"source_workbook": str(workbook_path)},
            }

        process_name = f"{parent_code} cathode parent process"
        process_runs.append(
            {
                "name": process_name,
                "process_type": ProcessType.COATING,
                "protocol_key": protocol_key,
                "started_at": parent["build_date"],
                "completed_at": None,
                "settings_json": {
                    **parent["safe_metrics"],
                    "source_workbook": str(workbook_path),
                    "source_sheet": parent["source_sheet"],
                    "active_material_source": parent["active_material_source"],
                    "mixing_strategy": mixing_strategy,
                    "current_collector": current_collector,
                    "separator": separator,
                },
                "notes": _combine_notes(parent["notes"], parent["result_notes"]),
            }
        )

        parent_formulation = _apply_batch_material_hints(
            [_serialize_component(component) for component in parent["formulation"]],
            active_material_source=parent["active_material_source"],
        )
        parent_metadata = {
            "source_workbook": str(workbook_path),
            "source_sheet": parent["source_sheet"],
            "active_material_source": parent["active_material_source"],
            "branch_descriptors": parent["branch_descriptors"],
            "safe_metrics": parent["safe_metrics"],
            "current_collector": current_collector,
            "separator": separator,
        }
        if stable_parent_loading:
            parent_metadata["loading_backfill"] = stable_parent_loading
        if exact_summary:
            parent_metadata["summary_row"] = exact_summary["raw_name"]
            parent_metadata["default_electrolyte"] = exact_summary.get("electrolyte")
            parent_metadata["default_cycling_strategy"] = exact_summary.get("cycling_strategy")
            parent_metadata["default_calendared"] = exact_summary.get("calendared")
            if exact_summary.get("comparison_ref"):
                parent_metadata["comparison_ref"] = exact_summary["comparison_ref"]
            if exact_summary.get("comparison_carbon_material"):
                parent_metadata["comparison_carbon_material"] = exact_summary["comparison_carbon_material"]

        if parent_code not in OMITTED_PARENT_BATCHES:
            parent_batches.append(
                {
                    "batch_name": parent_code,
                    "parent_batch_name": None,
                    "process_run_name": process_name,
                    "active_material_name": "NMC811",
                    "formulation_json": parent_formulation,
                    "notes": _combine_notes(parent["notes"], exact_summary.get("notes") if exact_summary else None, exact_legacy.get("notes") if exact_legacy else None),
                    "metadata_json": parent_metadata,
                    "source_kind": "parent_sheet",
                    "source_created_at": _as_datetime_from_date(parent["build_date"]),
                    "legacy_experiment_ids": [record["id"] for record in exact_legacy_records],
                }
            )
        else:
            notes.append(
                f"{parent_code} is not persisted as a shared parent batch because its child branches are independent electrodes."
            )

        if exact_summary and exact_legacy:
            mismatch_items.extend(
                _compare_summary_to_legacy(
                    record_name=parent_code,
                    scope="parent_batch",
                    summary_row=exact_summary,
                    legacy_record=exact_legacy,
                    expected_formulation=parent_formulation,
                )
            )

        if exact_summary and not exact_legacy and not any(record["match_key"] != parent_code for record in legacy_rows):
            if parent_code in EXPECTED_PARENT_ONLY_BATCHES:
                notes.append(
                    f"{parent_code} is treated as a user-confirmed parent-only ontology batch with no standalone legacy experiment."
                )
            else:
                mismatch_items.append(
                    _create_issue(
                        status="missing_in_legacy",
                        scope="parent_batch",
                        record_name=parent_code,
                        field="legacy_match",
                        message="Parent batch appears in the workbook summary but has no exact legacy experiment match.",
                        workbook_ref=exact_summary["workbook_ref"],
                    )
                )

        if exact_legacy:
            experiment_links.append(
                {
                    "source_batch_name": parent_code,
                    "legacy_experiment_id": exact_legacy["id"],
                    "relationship_type": "evaluated_in",
                    "notes": "Exact parent match between workbook summary and legacy experiment.",
                }
            )
            cell_builds.extend(
                _create_cell_build_specs(
                    source_batch_name=parent_code,
                    legacy_record=exact_legacy,
                    relationship_type="built_into",
                )
            )

            pressed = parent["safe_metrics"].get("pressed_thickness_um")
            if (
                pressed is not None
                and exact_legacy.get("pressed_thickness_um") is not None
                and not _numbers_close(pressed, exact_legacy["pressed_thickness_um"])
            ):
                mismatch_items.append(
                    _create_issue(
                        status="mismatch",
                        scope="parent_batch",
                        record_name=parent_code,
                        field="pressed_thickness_um",
                        message="Sheet pressed thickness does not match the existing database entry.",
                        workbook_value=pressed,
                        legacy_value=exact_legacy["pressed_thickness_um"],
                        legacy_ref=f"cell_experiments:{exact_legacy['id']}",
                    )
                )

        if exact_legacy_records and len(exact_legacy_records) > 1:
            notes.append(
                f"{parent_code} has {len(exact_legacy_records)} legacy exact matches; the importer will prefer repaired rows when present."
            )

        non_parent_summary_rows = [row for row in summary_rows if not row["is_exact_parent"]]
        legacy_variant_groups: Dict[str, List[Dict[str, Any]]] = {}
        for record in legacy_rows:
            if record["match_key"] != parent_code:
                legacy_variant_groups.setdefault(record["match_key"], []).append(record)

        child_variant_keys = {row["match_key"] for row in non_parent_summary_rows if row["match_key"]}
        child_variant_keys.update(legacy_variant_groups.keys())

        for variant_key in sorted(child_variant_keys):
            summary_row = next((row for row in non_parent_summary_rows if row["match_key"] == variant_key), None)
            related_legacy = legacy_variant_groups.get(variant_key, [])
            primary_legacy = _choose_primary_legacy_record(related_legacy)
            variant_name = summary_row["raw_name"] if summary_row else (primary_legacy["cell_name"] if primary_legacy else variant_key)
            variant_override = USER_VARIANT_OVERRIDES.get(variant_key, {})
            descriptor = (
                variant_override.get("variant_descriptor")
                or parent["branch_descriptors"].get(variant_key)
                or (summary_row.get("comparison_carbon_material") if summary_row else None)
            )
            variant_formulation = _build_variant_formulation(
                parent["formulation"],
                descriptor=descriptor,
                legacy_override=primary_legacy.get("formulation") if primary_legacy and not summary_row else None,
                user_override=variant_override,
            )
            variant_formulation = _apply_batch_material_hints(
                variant_formulation,
                active_material_source=parent["active_material_source"],
            )
            variant_notes = _combine_notes(
                descriptor,
                summary_row.get("notes") if summary_row else None,
                primary_legacy.get("notes") if primary_legacy else None,
                variant_override.get("override_notes"),
            )
            stable_variant_loading = _derive_stable_legacy_loading(related_legacy)
            is_parentless_variant = variant_key in PARENTLESS_VARIANT_KEYS
            variant_metadata = {
                "source_workbook": str(workbook_path),
                "parent_batch_name": None if is_parentless_variant else parent_code,
                "active_material_source": parent["active_material_source"],
                "summary_row": summary_row["raw_name"] if summary_row else None,
                "legacy_aliases": [record["cell_name"] for record in related_legacy],
                "electrolyte": summary_row.get("electrolyte") if summary_row else (primary_legacy.get("electrolyte") if primary_legacy else None),
                "cycler_channels": summary_row.get("cycler_channels") if summary_row else None,
                "cycling_strategy": summary_row.get("cycling_strategy") if summary_row else None,
                "calendared": (
                    summary_row.get("calendared")
                    if summary_row and summary_row.get("calendared") is not None
                    else (primary_legacy.get("calendared") if primary_legacy else None)
                ),
                "variant_descriptor": descriptor,
                "comparison_ref": summary_row.get("comparison_ref") if summary_row else None,
                "comparison_carbon_material": summary_row.get("comparison_carbon_material") if summary_row else None,
                "user_override_applied": bool(variant_override),
            }
            if is_parentless_variant:
                variant_metadata["detached_parent_sheet"] = parent_code
            if stable_variant_loading:
                variant_metadata["loading_mg"] = stable_variant_loading["value"]
                variant_metadata["loading_backfill"] = stable_variant_loading
            variant_metadata["override_metadata"] = _merge_metadata_dicts(variant_metadata.get("override_metadata"), variant_override.get("metadata"))
            if variant_metadata["override_metadata"] is None:
                variant_metadata.pop("override_metadata")

            variant_batches.append(
                {
                    "batch_name": variant_name,
                    "parent_batch_name": None if is_parentless_variant else parent_code,
                    "process_run_name": process_name,
                    "active_material_name": "NMC811",
                    "formulation_json": variant_formulation,
                    "notes": variant_notes,
                    "metadata_json": variant_metadata,
                    "source_kind": (
                        "independent_root_variant"
                        if is_parentless_variant
                        else ("merged" if summary_row and primary_legacy else ("summary_only" if summary_row else "legacy_only"))
                    ),
                    "source_created_at": _as_datetime_from_date(parent["build_date"] or (summary_row.get("date") if summary_row else None)),
                    "legacy_experiment_ids": [record["id"] for record in related_legacy],
                }
            )

            if summary_row and primary_legacy:
                mismatch_items.extend(
                    _compare_summary_to_legacy(
                        record_name=variant_name,
                        scope="variant_batch",
                        summary_row=summary_row,
                        legacy_record=primary_legacy,
                        expected_formulation=variant_formulation,
                    )
                )

            if summary_row and not primary_legacy:
                if variant_key in EXPECTED_WORKBOOK_ONLY_VARIANTS:
                    notes.append(
                        f"{variant_name} is treated as a workbook-only ontology branch with no required legacy experiment row."
                    )
                else:
                    mismatch_items.append(
                        _create_issue(
                            status="missing_in_legacy",
                            scope="variant_batch",
                            record_name=variant_name,
                            field="legacy_match",
                            message="Workbook variant has no matching legacy experiment entry.",
                            workbook_ref=summary_row["workbook_ref"],
                        )
                    )

            has_explicit_variant_definition = bool(
                parent["branch_descriptors"].get(variant_key) or USER_VARIANT_OVERRIDES.get(variant_key)
            )
            if primary_legacy and not summary_row and not has_explicit_variant_definition:
                mismatch_items.append(
                    _create_issue(
                        status="legacy_only",
                        scope="variant_batch",
                        record_name=variant_name,
                        field="workbook_match",
                        message="Legacy experiment exists without a matching summary-sheet variant row.",
                        legacy_ref=f"cell_experiments:{primary_legacy['id']}",
                    )
                )

            for legacy_record in related_legacy:
                experiment_links.append(
                    {
                        "source_batch_name": variant_name,
                        "legacy_experiment_id": legacy_record["id"],
                        "relationship_type": "evaluated_in",
                        "notes": "Variant branch linked to legacy experiment record.",
                    }
                )
            if primary_legacy:
                cell_builds.extend(
                    _create_cell_build_specs(
                        source_batch_name=variant_name,
                        legacy_record=primary_legacy,
                        relationship_type="built_into",
                    )
                )

    for full_cell_name, override in FULL_CELL_OVERRIDES.items():
        legacy_record = legacy["full_cells_by_name"].get(full_cell_name)
        if not legacy_record:
            mismatch_items.append(
                _create_issue(
                    status="legacy_only",
                    scope="full_cell",
                    record_name=full_cell_name,
                    field="override_target",
                    message="Configured full-cell cathode mapping could not be resolved in the legacy database.",
                )
            )
            continue
        cell_builds.extend(
            _create_cell_build_specs(
                source_batch_name=override["source_batch_name"],
                legacy_record=legacy_record,
                relationship_type="cathode_source_for",
                relationship_notes=override["notes"],
            )
        )

    counts = {
        "materials": len(materials),
        "protocol_versions": len(protocols),
        "process_runs": len(process_runs),
        "parent_batches": len(parent_batches),
        "variant_batches": len(variant_batches),
        "legacy_experiment_links": len(experiment_links),
        "cell_builds": len(cell_builds),
        "review_items": len(review_items),
        "mismatch_items": len(mismatch_items),
    }

    return {
        "source_workbook": str(workbook_path),
        "legacy_db_path": str(legacy_db_path),
        "comparison_workbook_path": str(comparison_workbook_path) if comparison_workbook_path else None,
        "counts": counts,
        "materials": sorted(materials.values(), key=lambda item: item["name"]),
        "protocol_versions": sorted(protocols.values(), key=lambda item: item["name"]),
        "process_runs": process_runs,
        "parent_batches": sorted(parent_batches, key=lambda item: item["batch_name"]),
        "variant_batches": sorted(variant_batches, key=lambda item: item["batch_name"]),
        "legacy_experiment_links": experiment_links,
        "cell_builds": sorted(cell_builds, key=lambda item: item["build_name"]),
        "review_items": review_items,
        "mismatch_items": mismatch_items,
        "notes": notes,
    }


async def _get_by_unique_field(
    db: AsyncSession,
    model: Any,
    field_name: str,
    value: Any,
) -> Optional[Any]:
    result = await db.execute(select(model).where(getattr(model, field_name) == value))
    return result.scalars().first()


def _merge_json(existing: Optional[Dict[str, Any]], incoming: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not existing and not incoming:
        return None
    merged = dict(existing or {})
    merged.update(incoming or {})
    return merged


async def _upsert_material(db: AsyncSession, spec: Dict[str, Any]) -> Material:
    material = await _get_by_unique_field(db, Material, "name", spec["name"])
    if material is None:
        material = Material(**spec)
        db.add(material)
        await db.flush()
        return material

    material.category = spec["category"]
    material.manufacturer = spec.get("manufacturer") or material.manufacturer
    material.description = spec.get("description") or material.description
    material.metadata_json = _merge_json(material.metadata_json, spec.get("metadata_json"))
    await db.flush()
    return material


async def _upsert_protocol_version(db: AsyncSession, spec: Dict[str, Any]) -> ProtocolVersion:
    result = await db.execute(
        select(ProtocolVersion)
        .where(ProtocolVersion.name == spec["name"])
        .where(ProtocolVersion.version == spec["version"])
    )
    protocol = result.scalars().first()
    if protocol is None:
        protocol = ProtocolVersion(**spec)
        db.add(protocol)
        await db.flush()
        return protocol

    protocol.protocol_type = spec["protocol_type"]
    protocol.description = spec.get("description") or protocol.description
    protocol.step_definition_json = spec.get("step_definition_json") or protocol.step_definition_json
    protocol.metadata_json = _merge_json(protocol.metadata_json, spec.get("metadata_json"))
    await db.flush()
    return protocol


async def _upsert_process_run(
    db: AsyncSession,
    spec: Dict[str, Any],
    protocol_version_id: Optional[int],
) -> ProcessRun:
    process_run = await _get_by_unique_field(db, ProcessRun, "name", spec["name"])
    payload = {
        "name": spec["name"],
        "process_type": spec["process_type"],
        "protocol_version_id": protocol_version_id,
        "started_at": spec.get("started_at"),
        "completed_at": spec.get("completed_at"),
        "settings_json": spec.get("settings_json"),
        "notes": spec.get("notes"),
    }
    if process_run is None:
        process_run = ProcessRun(**payload)
        db.add(process_run)
        await db.flush()
        return process_run

    process_run.process_type = spec["process_type"]
    process_run.protocol_version_id = protocol_version_id
    process_run.started_at = spec.get("started_at") or process_run.started_at
    process_run.completed_at = spec.get("completed_at") or process_run.completed_at
    process_run.settings_json = _merge_json(process_run.settings_json, spec.get("settings_json"))
    process_run.notes = spec.get("notes") or process_run.notes
    await db.flush()
    return process_run


async def _upsert_electrode_batch(
    db: AsyncSession,
    spec: Dict[str, Any],
    *,
    process_run_id: Optional[int],
    active_material_id: Optional[int],
) -> ElectrodeBatch:
    batch = await _get_by_unique_field(db, ElectrodeBatch, "batch_name", spec["batch_name"])
    payload = {
        "batch_name": spec["batch_name"],
        "electrode_role": ElectrodeRole.CATHODE,
        "active_material_id": active_material_id,
        "process_run_id": process_run_id,
        "formulation_json": spec.get("formulation_json"),
        "notes": spec.get("notes"),
        "metadata_json": spec.get("metadata_json"),
        "created_at": spec.get("source_created_at"),
    }
    if batch is None:
        batch = ElectrodeBatch(**payload)
        db.add(batch)
        await db.flush()
        return batch

    batch.active_material_id = active_material_id
    batch.process_run_id = process_run_id
    batch.formulation_json = spec.get("formulation_json") or batch.formulation_json
    batch.notes = spec.get("notes") or batch.notes
    batch.metadata_json = _merge_json(batch.metadata_json, spec.get("metadata_json"))
    if spec.get("source_created_at") is not None:
        batch.created_at = spec["source_created_at"]
    await db.flush()
    return batch


async def _upsert_cell_build(db: AsyncSession, spec: Dict[str, Any]) -> CellBuild:
    source_created_at = _as_datetime_from_date(spec.get("build_date"))
    build = await _get_by_unique_field(db, CellBuild, "build_name", spec["build_name"])
    payload = {
        "build_name": spec["build_name"],
        "chemistry": spec.get("chemistry"),
        "form_factor": spec.get("form_factor"),
        "build_date": spec.get("build_date"),
        "status": spec.get("status", CellBuildStatus.PLANNED),
        "notes": spec.get("notes"),
        "legacy_project_id": spec.get("legacy_project_id"),
        "legacy_experiment_id": spec.get("legacy_experiment_id"),
        "legacy_cell_id": spec.get("legacy_cell_id"),
        "legacy_test_number": spec.get("legacy_test_number"),
        "metadata_json": spec.get("metadata_json"),
        "created_at": source_created_at,
    }
    if build is None:
        build = CellBuild(**payload)
        db.add(build)
        await db.flush()
        return build

    build.chemistry = spec.get("chemistry") or build.chemistry
    build.form_factor = spec.get("form_factor") or build.form_factor
    build.build_date = spec.get("build_date") or build.build_date
    build.status = spec.get("status", build.status)
    build.notes = spec.get("notes") or build.notes
    build.legacy_project_id = spec.get("legacy_project_id") or build.legacy_project_id
    build.legacy_experiment_id = spec.get("legacy_experiment_id") or build.legacy_experiment_id
    build.legacy_cell_id = spec.get("legacy_cell_id") or build.legacy_cell_id
    build.legacy_test_number = spec.get("legacy_test_number") or build.legacy_test_number
    build.metadata_json = _merge_json(build.metadata_json, spec.get("metadata_json"))
    if source_created_at is not None:
        build.created_at = source_created_at
    await db.flush()
    return build


async def _upsert_lineage_edge(
    db: AsyncSession,
    *,
    parent_type: LineageEntityType,
    parent_id: int,
    child_type: LineageEntityType,
    child_id: int,
    relationship_type: str,
    source: str,
    confidence: Optional[float] = None,
    notes: Optional[str] = None,
) -> LineageEdge:
    result = await db.execute(
        select(LineageEdge)
        .where(LineageEdge.parent_type == parent_type)
        .where(LineageEdge.parent_id == parent_id)
        .where(LineageEdge.child_type == child_type)
        .where(LineageEdge.child_id == child_id)
        .where(LineageEdge.relationship_type == relationship_type)
    )
    edge = result.scalars().first()
    if edge is None:
        edge = LineageEdge(
            parent_type=parent_type,
            parent_id=parent_id,
            child_type=child_type,
            child_id=child_id,
            relationship_type=relationship_type,
            source=source,
            confidence=confidence,
            notes=notes,
        )
        db.add(edge)
        await db.flush()
        return edge

    edge.source = source
    edge.confidence = confidence if confidence is not None else edge.confidence
    edge.notes = notes or edge.notes
    await db.flush()
    return edge


async def _remove_batch_and_edges(db: AsyncSession, batch_name: str) -> None:
    batch = await _get_by_unique_field(db, ElectrodeBatch, "batch_name", batch_name)
    if batch is None:
        return

    await db.execute(
        delete(LineageEdge).where(
            or_(
                and_(
                    LineageEdge.parent_type == LineageEntityType.ELECTRODE_BATCH,
                    LineageEdge.parent_id == batch.id,
                ),
                and_(
                    LineageEdge.child_type == LineageEntityType.ELECTRODE_BATCH,
                    LineageEdge.child_id == batch.id,
                ),
            )
        )
    )
    await db.delete(batch)
    await db.flush()


async def import_nmc_cathode_workbook(
    db: AsyncSession,
    *,
    workbook_path: str,
    legacy_db_path: Optional[str] = None,
    comparison_workbook_path: Optional[str] = None,
    export_dir: Optional[str] = None,
    persist: bool = False,
) -> Dict[str, Any]:
    workbook = Path(workbook_path).expanduser().resolve()
    legacy_db = Path(legacy_db_path).expanduser().resolve() if legacy_db_path else DEFAULT_LEGACY_DB_PATH.resolve()
    comparison_workbook = (
        Path(comparison_workbook_path).expanduser().resolve()
        if comparison_workbook_path
        else None
    )
    review_export_dir = Path(export_dir).expanduser().resolve() if export_dir else None
    if not workbook.exists():
        raise ValueError(f"Workbook not found: {workbook}")
    if not legacy_db.exists():
        raise ValueError(f"Legacy database not found: {legacy_db}")
    if comparison_workbook is not None and not comparison_workbook.exists():
        raise ValueError(f"Comparison workbook not found: {comparison_workbook}")

    plan = _prepare_import_plan(workbook, legacy_db, comparison_workbook)
    plan["persisted"] = False
    if review_export_dir is not None:
        plan["review_export"] = _write_review_export(plan, review_export_dir)

    if not persist:
        return plan

    material_map: Dict[str, Material] = {}
    for material_spec in plan["materials"]:
        material = await _upsert_material(db, material_spec)
        material_map[material.name] = material

    protocol_map: Dict[Tuple[str, str], ProtocolVersion] = {}
    for protocol_spec in plan["protocol_versions"]:
        protocol = await _upsert_protocol_version(db, protocol_spec)
        protocol_map[(protocol.name, protocol.version)] = protocol

    process_map: Dict[str, ProcessRun] = {}
    for process_spec in plan["process_runs"]:
        protocol_key = process_spec.get("protocol_key")
        protocol = protocol_map.get(protocol_key) if protocol_key else None
        process_run = await _upsert_process_run(db, process_spec, protocol.id if protocol else None)
        process_map[process_run.name] = process_run

    batch_map: Dict[str, ElectrodeBatch] = {}
    for batch_spec in [*plan["parent_batches"], *plan["variant_batches"]]:
        process_run = process_map.get(batch_spec["process_run_name"])
        active_material = material_map.get(batch_spec["active_material_name"])
        batch = await _upsert_electrode_batch(
            db,
            batch_spec,
            process_run_id=process_run.id if process_run else None,
            active_material_id=active_material.id if active_material else None,
        )
        batch_map[batch.batch_name] = batch

    for batch_spec in [*plan["parent_batches"], *plan["variant_batches"]]:
        batch = batch_map[batch_spec["batch_name"]]
        parent_batch_name = batch_spec.get("parent_batch_name")
        if parent_batch_name:
            parent_batch = batch_map[parent_batch_name]
            await _upsert_lineage_edge(
                db,
                parent_type=LineageEntityType.ELECTRODE_BATCH,
                parent_id=parent_batch.id,
                child_type=LineageEntityType.ELECTRODE_BATCH,
                child_id=batch.id,
                relationship_type="branches_to",
                source=IMPORT_SOURCE,
                confidence=1.0,
                notes="Workbook parent/branch relationship.",
            )
        for component in batch_spec.get("formulation_json") or []:
            material = material_map.get(component.get("Component"))
            if not material:
                continue
            await _upsert_lineage_edge(
                db,
                parent_type=LineageEntityType.MATERIAL,
                parent_id=material.id,
                child_type=LineageEntityType.ELECTRODE_BATCH,
                child_id=batch.id,
                relationship_type="formulates_into",
                source=IMPORT_SOURCE,
                confidence=1.0,
                notes=f"Imported formulation component for {batch.batch_name}.",
            )

    for link_spec in plan["legacy_experiment_links"]:
        batch = batch_map.get(link_spec["source_batch_name"])
        if not batch:
            continue
        await _upsert_lineage_edge(
            db,
            parent_type=LineageEntityType.ELECTRODE_BATCH,
            parent_id=batch.id,
            child_type=LineageEntityType.EXPERIMENT,
            child_id=link_spec["legacy_experiment_id"],
            relationship_type=link_spec["relationship_type"],
            source=IMPORT_SOURCE,
            confidence=0.9,
            notes=link_spec.get("notes"),
        )

    for build_spec in plan["cell_builds"]:
        batch = batch_map.get(build_spec["source_batch_name"])
        if not batch:
            continue
        build = await _upsert_cell_build(db, build_spec)
        await _upsert_lineage_edge(
            db,
            parent_type=LineageEntityType.ELECTRODE_BATCH,
            parent_id=batch.id,
            child_type=LineageEntityType.CELL_BUILD,
            child_id=build.id,
            relationship_type=build_spec["relationship_type"],
            source=IMPORT_SOURCE,
            confidence=0.9,
            notes=build_spec.get("relationship_notes"),
        )

    for omitted_batch_name in OMITTED_PARENT_BATCHES:
        if omitted_batch_name not in batch_map:
            await _remove_batch_and_edges(db, omitted_batch_name)

    plan["persisted"] = True
    return plan
